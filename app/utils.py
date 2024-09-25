""" Set generel functions """

import base64
import requests
import msal

from io import BytesIO
from xhtml2pdf import pisa
from xhtml2pdf.files import pisaFileObject
from openpyxl import load_workbook, Workbook, styles
from datetime import timedelta, datetime
from collections import defaultdict

from django.conf import settings
from django.template.loader import get_template
from django.utils.http import urlsafe_base64_encode
from django.utils.encoding import force_bytes
from django.contrib.sites.shortcuts import get_current_site
from django.http import HttpResponse, HttpResponseNotFound
from django.urls import reverse, NoReverseMatch
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Q, Case, When, Value, IntegerField, Count
from django.core.mail import send_mail, get_connection, EmailMessage
from django.template.loader import render_to_string

from db_logger.utils import create_log
from .models import *
from .choices import *
from .tokens import account_activation_token
from .api.dhl import dhl_check_status, dhl_check_bulk_status
from .api.go_express import go_express_check_status
from .api.brevo import brevo_send_order_shipped, brevo_send_activate_staff_user, brevo_send_order_confirmation,\
                        brevo_send_activate_user, brevo_send_reset_password, brevo_send_order_cancelled, brevo_send_pre_invoice,\
                        brevo_send_invoice, brevo_send_new_order_created, brevo_order_ready_for_pickup

# Excel Datei Spalten und Zeilenweise auslesen
#pylint: disable=too-many-arguments
def ws_lookup_neighbour(worksheet, min_row, max_col, max_row, value, offset=(0,0)):
    """ Find values by header """
    for row in worksheet.iter_rows(min_row = min_row, max_col = max_col, max_row = max_row):
        for cell in row:
            if cell.value == value:
                return cell.offset(*offset).value
        return None

def custom_currency_format(amount):
    """ Format float to euro """
    formatted_amount = f"{amount:,.2f}".replace(",", ";").replace(".", ",").replace(";", ".")
    return f"{formatted_amount} €"

def remove_product_from_cart(order_product_id):
    """ Remove product from cart by order_product_id """

    data = {
        'deleted': False
    }

    try:
        OrderProducts.objects.get(id=order_product_id).delete()

        data['deleted'] = True

    except Products.DoesNotExist:

        data['product_not_found'] = True

    except ObjectDoesNotExist:

        data['object_not_exist'] = True

    return data

def get_full_domain(request):
    """ Get Full domain by request """

    full_domain = request.build_absolute_uri().split(str(get_current_site(request)))[0] + str(get_current_site(request))

    return full_domain

def generate_invoice_insurance(invoice_id):
    """ Generate invoice """

    main_settings = MainSettings.objects.first()
    invoice = Invoices.objects.get(id=invoice_id)
    invoice_items = InvoiceItems.objects.filter(invoice=invoice)

    response = None

    if main_settings and main_settings.invoice_layout:

        template = get_template('app/exports/invoice_insurance.html')

        layout = main_settings.invoice_layout.path
        font_family_regular = os.path.join(settings.FONTS_DIR, 'HelveticaNowText-Regular' + '.ttf')
        font_family_bold = os.path.join(settings.FONTS_DIR, 'HelveticaNowText-Bold' + '.ttf')

        context = {
            'layout': layout,
            'main_settings': main_settings,
            'invoice': invoice,
            'invoice_items': invoice_items,
            'font_family_regular': font_family_regular,
            'font_family_bold': font_family_bold,
        }

        filename = 'Rechnung.pdf'

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="%s"'%(filename)
        html = template.render(context)

        pisaFileObject.getNamedFile = lambda self: self.uri

        #pylint: disable=unused-variable
        pisa_status = pisa.CreatePDF(html, dest=response)

    return response

def generate_invoice_customer(invoice_id, invoice_type):
    """ Generate invoice """

    main_settings = MainSettings.objects.first()
    invoice = Invoices.objects.get(id=invoice_id)
    invoice_items = InvoiceItems.objects.filter(invoice=invoice)

    if main_settings and main_settings.invoice_layout:

        template = get_template('app/exports/invoice_customer.html')

        layout = main_settings.invoice_layout.path
        font_family_light = os.path.join(settings.FONTS_DIR, 'HelveticaNowText-Light.ttf')
        font_family_regular = os.path.join(settings.FONTS_DIR, 'HelveticaNowText-Regular.ttf')
        font_family_bold = os.path.join(settings.FONTS_DIR, 'HelveticaNowText-Bold.ttf')

        context = {
            'layout': layout,
            'main_settings': main_settings,
            'invoice': invoice,
            'invoice_items': invoice_items,
            'font_family_light': font_family_light,
            'font_family_regular': font_family_regular,
            'font_family_bold': font_family_bold,
        }

        filename = 'Rechnung.pdf'

        html = template.render(context)

        pisaFileObject.getNamedFile = lambda self: self.uri

        if invoice_type == 'http':

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="%s"'%(filename)

            #pylint: disable=unused-variable
            pisa_status = pisa.CreatePDF(html, dest=response)

        elif invoice_type == 'base':

            # Redirect PDF output to a BytesIO stream
            output = BytesIO()
            pisa_status = pisa.CreatePDF(html, dest=output)

            # Encode the BytesIO stream to Base64
            response = base64.b64encode(output.getvalue()).decode('utf-8')

    else:
        response = HttpResponseNotFound('Rechnung konnte nicht generiert werden!')

    return response

def get_choice_value_by_display(choices, display_name):
    """ Get choices value by display name """
    for value, display in choices:
        if display == display_name:
            return value
    return None

def get_values_by_display_nams(choices, display_names):
    """ Get all values of a choices list by display_names """

    values = []

    for display_name in display_names:
        for value, name in choices:
            if name == display_name:
                values.append(value)
                break

    return values 

# def create_log(reference, message, category='info', user='Unknown', stack_trace=''):
#     """ create custom log entry """

#     log_item = Logger.objects.create(user=user, category=category, reference=reference, message=message, stack_trace=stack_trace)

#     if log_item.category in ['error', 'fatal']:
#         send_error_mail(log_item)

#     return True

def create_product_stock_items():
    """ Create product stock items """

    for pharmacy in Pharmacies.objects.all():
        for product in Products.objects.all():
            if not StockProducts.objects.filter(product=product, pharmacy=pharmacy).exists():
                StockProducts.objects.create(product=product, pharmacy=pharmacy)

######### Import functions #########
def import_terpene(import_file, request):
    """ Import terpene """

    workbook = load_workbook(import_file, data_only=True)

    ##### Natürliches Vorkommen #####
    worksheet = workbook['NaturalOccurrence']

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Name', (row_offset - 2, 0))

        try:
            item = NaturalOccurrence.objects.get(name=name)

        except ObjectDoesNotExist:

            item, created = NaturalOccurrence.objects.get_or_create(id=item_id, defaults={'name': name})

            if not created:
                item.name = name.strip()
                item.save()

        row_offset += 1

    ##### Aromen #####
    worksheet = workbook['Flavors']

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Name', (row_offset - 2, 0))

        try:
            item = Flavors.objects.get(name=name)

        except ObjectDoesNotExist:

            item, created = Flavors.objects.get_or_create(id=item_id, defaults={'name': name})

            if not created:
                item.name = name.strip()
                item.save()

        row_offset += 1

    ##### Terpenwirkungen ######
    worksheet = workbook['TerpeneEffects']

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Name', (row_offset - 2, 0))

        try:
            item = TerpeneEffects.objects.get(name=name)

        except ObjectDoesNotExist:

            item, created = TerpeneEffects.objects.get_or_create(id=item_id, defaults={'name': name})

            if not created:
                item.name = name.strip()
                item.save()

        row_offset += 1

    ##### Indikationen #####
    worksheet = workbook['Indications']

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Name', (row_offset - 2, 0))

        try:
            item = Indications.objects.get(name=name)

        except ObjectDoesNotExist:

            item, created = Indications.objects.get_or_create(id=item_id, defaults={'name': name})

            if not created:
                item.name = name.strip()
                item.save()

        row_offset += 1

    ##### Terpene #####
    worksheet = workbook['Terpene']

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Name', (row_offset - 2, 0))
        natural_occurrence = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Natürliches Vorkommen', (row_offset - 2, 0))
        flavors = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Aromen', (row_offset - 2, 0))
        terpene_effect = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Terpenwirkungen', (row_offset - 2, 0))
        indications = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Indikationen', (row_offset - 2, 0))
        description = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Beschreibung', (row_offset - 2, 0))

        terpene = None

        if item_id and item_id != '':

            try:
                terpene = Terpene.objects.get(id=item_id)
            except ObjectDoesNotExist:
                pass

        if not terpene:

            terpene, created = Terpene.objects.update_or_create(name=name, defaults={'description': description})

        # Natürliche vorkommen
        terpene.natural_occurrence.clear()
        for name in natural_occurrence.split(';'):
            try:
                item = NaturalOccurrence.objects.get(name=name)
                terpene.natural_occurrence.add(item)
            except ObjectDoesNotExist:
                pass

        # Aromen
        terpene.flavors.clear()
        for name in flavors.split(';'):
            try:
                item = Flavors.objects.get(name=name)
                terpene.flavors.add(item)
            except ObjectDoesNotExist:
                pass

        # Terpenenwirkungen
        terpene.terpene_effect.clear()
        for name in terpene_effect.split(';'):
            try:
                item = TerpeneEffects.objects.get(name=name)
                terpene.terpene_effect.add(item)
            except ObjectDoesNotExist:
                pass

        # Indikationen
        terpene.indications.clear()
        for name in indications.split(';'):
            try:
                item = Indications.objects.get(name=name)
                terpene.indications.add(item)
            except ObjectDoesNotExist:
                pass

        terpene.save()

        row_offset += 1

    # Create log entry
    log_entry = {
        'reference': 'Dataimport',
        'message': 'Terpene Datei wurde importiert',
        'user': f'({ request.user.id }) { request.user.username }'
    }

    create_log(**log_entry)

def import_products(import_file, request=None):
    """ Import products """
    
    if import_file:
        workbook = load_workbook(import_file, data_only=True)
    else:
        products_file = get_products_file()
        workbook = load_workbook(products_file, data_only=True)

    # Kultivar
    worksheet = workbook['Cultivar']

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Name', (row_offset - 2, 0))

        if name:

            try:
                item = Cultivar.objects.get(name=name)

            except ObjectDoesNotExist:

                item, created = Cultivar.objects.get_or_create(id=item_id, defaults={'name': name})

                if not created:
                    item.name = name.strip()
                    item.save()

        row_offset += 1

    # Genetik
    worksheet = workbook['Genetics']

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Name', (row_offset - 2, 0))
        description = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Beschreibung', (row_offset - 2, 0))

        if name:

            try:
                item = Genetics.objects.get(name=name)

            except ObjectDoesNotExist:

                item, created = Genetics.objects.get_or_create(id=item_id, defaults={'name': name})

                if not created:
                    item.name = name.strip()

            item.description = description if description else ''
            item.save()

        row_offset += 1

    # Hersteller
    worksheet = workbook['Manufacturer']

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Name', (row_offset - 2, 0))
        description = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Beschreibung', (row_offset - 2, 0))

        if name:

            try:
                item = Manufacturer.objects.get(name=name)

            except ObjectDoesNotExist:

                item, created = Manufacturer.objects.get_or_create(id=item_id, defaults={'name': name})

                if not created:
                    item.name = name.strip()

            item.description = description if description else ''
            item.save()

        row_offset += 1

    # Herkunftsland
    worksheet = workbook['CountryOfOrigin']

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Name', (row_offset - 2, 0))

        try:
            item = CountryOfOrigin.objects.get(name=name)

        except ObjectDoesNotExist:

            item, created = CountryOfOrigin.objects.get_or_create(id=item_id, defaults={'name': name})

            if not created:
                item.name = name.strip()
                item.save()

        row_offset += 1

    # Lieferant
    worksheet = workbook['Supplier']

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Name', (row_offset - 2, 0))

        try:
            item = Supplier.objects.get(name=name)

        except ObjectDoesNotExist:

            item, created = Supplier.objects.get_or_create(id=item_id, defaults={'name': name})

            if not created:
                item.name = name.strip()
                item.save()

        row_offset += 1

    # PZN Auswahl
    worksheet = workbook['PZNChoices']

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        amount = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Menge', (row_offset - 2, 0))
        unit = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Einheit', (row_offset - 2, 0))

        try:
            item = PZNChoices.objects.get(amount=amount, unit=unit)

        except ObjectDoesNotExist:

            item, created = PZNChoices.objects.get_or_create(id=item_id, defaults={'amount': amount, 'unit': unit})

        row_offset += 1

    # Produkte
    worksheet = workbook['Products']

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        item_id = item_id.replace(' ', '') if item_id and isinstance(item_id, str) else item_id
        number = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Artikelnummer', (row_offset - 2, 0))
        number = number.replace(' ', '') if number and isinstance(number, str) else number
        name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Name', (row_offset - 2, 0))
        description = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Beschreibung', (row_offset - 2, 0))
        thc_value = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'THC', (row_offset - 2, 0))
        max_cbd_value = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'CBD', (row_offset - 2, 0))
        treatment = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Behandlung', (row_offset - 2, 0))
        supplier = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Lieferant', (row_offset - 2, 0))
        form = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Form', (row_offset - 2, 0))
        meta_title = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Meta-Titel', (row_offset - 2, 0))
        meta_description = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Meta-Description', (row_offset - 2, 0))

        if number and name:

            if item_id and item_id != 0:
                product = Products.objects.get(id=item_id)
            else:
                product, product_created = Products.objects.get_or_create(number=number)

            product.number = number
            product.name = name.strip()
            product.description = description.strip() if description else ''
            product.thc_value = float(str(thc_value).replace(',', '.')) / 100 if thc_value else 0
            product.max_cbd_value = float(str(max_cbd_value).replace(',', '.')) / 100 if max_cbd_value else 0
            product.treatment = get_choice_value_by_display(TreatmentChoices, treatment) if treatment else ''
            product.form = get_choice_value_by_display(CannabisFormChoices, form) if form else ''
            product.meta_title = meta_title if meta_title else ''
            product.meta_description = meta_description if meta_description else ''

            # Kultivar
            cultivar = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Kultivar', (row_offset - 2, 0))
            try:
                product.cultivar = Cultivar.objects.get(name=cultivar)
            except ObjectDoesNotExist:
                pass

            # Genetik
            genetics = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Genetik', (row_offset - 2, 0))
            try:
                product.genetics = Genetics.objects.get(name=genetics)
            except ObjectDoesNotExist:
                pass

            # Hersteller
            manufacturer = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Hersteller', (row_offset - 2, 0))
            try:
                product.manufacturer = Manufacturer.objects.get(name=manufacturer)
            except ObjectDoesNotExist:
                pass

            # Herkunftsland
            country_of_origin = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Herkunftsland', (row_offset - 2, 0))
            try:
                product.country_of_origin = CountryOfOrigin.objects.get(name=country_of_origin)
            except ObjectDoesNotExist:
                pass

            # Lieferant
            supplier = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Lieferant', (row_offset - 2, 0))
            try:
                product.supplier = Supplier.objects.get(name=supplier)
            except ObjectDoesNotExist:
                pass

            # Hauptterpene
            terpene = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Hauptterpene', (row_offset - 2, 0))

            if terpene:
                for item in terpene.split(';'):
                    try:
                        product.main_terpene.add(Terpene.objects.get(name=item))
                    except ObjectDoesNotExist:
                        pass

            product.save()

            # PZN Nummern
            for pzn_choice in PZNChoices.objects.all():

                pzn_title = f'PZN-Nummer ({pzn_choice.amount}{pzn_choice.get_unit_display()})'

                pzn_number = ws_lookup_neighbour(worksheet, 3, max_column, max_row, pzn_title, (row_offset - 2, 0))

                if pzn_number and pzn_number != '':
                    PZNAmounts.objects.update_or_create(number=pzn_number, defaults={'product': product, 'pzn_choice': pzn_choice})

            # Create product price if not exists
            for pharmacy in Pharmacies.objects.all():
                ProductPrices.objects.get_or_create(product=product, pharmacy=pharmacy)

        row_offset += 1

    # Create log entry
    log_entry = {
        'reference': 'Dataimport',
        'message': 'Produktstammdaten-Datei importiert',
        'user': f'({ request.user.id if request else "0" }) { request.user.username if request else "System" }'
    }

    create_log(**log_entry)

def import_product_prices(import_file, pharmacy, request):

    workbook = load_workbook(import_file, data_only=True)

    worksheet = workbook['ProductPrices']

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        item_id = item_id.replace(' ', '') if item_id and isinstance(item_id, str) else item_id
        number = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Artikelnummer', (row_offset - 2, 0))
        number = number.replace(' ', '') if number and isinstance(number, str) else number
        pirce_per_unit = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Preis pro Einheit', (row_offset - 2, 0))
        purchase_price = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Preis EK', (row_offset - 2, 0))
        selling_price = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Preis VK', (row_offset - 2, 0))
        status = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Status', (row_offset - 2, 0))
        active = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Aktiv', (row_offset - 2, 0))

        if number:
            try:
                if item_id and item_id != 0:
                    product = Products.objects.get(id=item_id)
                else:
                    product = Products.objects.get(number=number)
            except ObjectDoesNotExist:
                product = None

            if product:
                # Import prices
                pirce_per_unit = round(float(str(pirce_per_unit).replace(',', '.')), 2) if pirce_per_unit and pirce_per_unit != '-' else ''
                purchase_price = round(float(str(purchase_price).replace(',', '.')), 2) if purchase_price and purchase_price != '-' else 0
                selling_price = round(float(str(selling_price).replace(',', '.')), 2) if selling_price and selling_price != '-' else 0
                status = status if status and purchase_price and purchase_price != 0 else 0
                active = True if str(active) == '1' else False

                # Create product price if not exists
                product_price, created = ProductPrices.objects.get_or_create(product=product, pharmacy=pharmacy)
                product_price.price_per_unit = pirce_per_unit
                product_price.purchase_price = purchase_price
                product_price.self_payer_selling_price = selling_price
                product_price.status = status
                product_price.active = active
                product_price.save()

        row_offset += 1
        
    # Create log entry
    log_entry = {
        'reference': 'Dataimport',
        'message': 'Produktstammdaten-Datei importiert',
        'user': f'({ request.user.id if request else "0" }) { request.user.username if request else "System" }'
    }

    create_log(**log_entry)

def import_faqs(import_file, request):
    """ Import faqs """

    workbook = load_workbook(import_file, data_only=True)

    for group in FAQGroups.objects.all():

        try:
            worksheet = workbook[group.name]

            max_row = worksheet.max_row
            max_column = worksheet.max_column
            row_offset = 3

            while row_offset < max_row:
                question = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Frage', (row_offset - 2, 0))
                answer = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Antwort', (row_offset - 2, 0))

                if question and answer:
                    faq, created = FAQs.objects.get_or_create(question=question, group=group, defaults={'answer': answer})

                    if not created:
                        faq.save()

                row_offset += 1

        except Exception:
            pass

    # Create log entry
    log_entry = {
        'reference': 'Dataimport',
        'message': 'FAQs-Datei importiert',
        'user': f'({ request.user.id }) { request.user.username }'
    }
    create_log(**log_entry)

def import_blog(import_file, request):
    """ Import blog """

    workbook = load_workbook(import_file, data_only=True)

    worksheet = workbook.worksheets[0]

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 1

    while row_offset < max_row:
        title = ws_lookup_neighbour(worksheet, 1, max_column, max_row, 'Themen', (row_offset, 0))
        teaser = ws_lookup_neighbour(worksheet, 1, max_column, max_row, 'Teasertext', (row_offset, 0))
        text = ws_lookup_neighbour(worksheet, 1, max_column, max_row, 'Textinhalt', (row_offset, 0))
        source = ws_lookup_neighbour(worksheet, 1, max_column, max_row, 'Quellen', (row_offset, 0))
        
        if title:
            CannabisBlog.objects.create(title=title, teaser=teaser, text=text, source=source)

        row_offset += 1

    # Create log entry
    log_entry = {
        'reference': 'Dataimport',
        'message': 'Blog-Datei importiert',
        'user': f'({ request.user.id }) { request.user.username }'
    }
    create_log(**log_entry)

def import_lexicon(import_file, request):
    """ Import lexicon """

    workbook = load_workbook(import_file, data_only=True)

    worksheet = workbook.worksheets[0]

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 1

    while row_offset < max_row:
        title = ws_lookup_neighbour(worksheet, 1, max_column, max_row, 'Titel', (row_offset, 0))
        description = ws_lookup_neighbour(worksheet, 1, max_column, max_row, 'Beschreibung', (row_offset, 0))
        
        if title:
            Lexicon.objects.create(title=title, description=description)

        row_offset += 1

    # Create log entry
    log_entry = {
        'reference': 'Dataimport',
        'message': 'Lexikon-Datei importiert',
        'user': f'({ request.user.id }) { request.user.username }'
    }
    create_log(**log_entry)

def import_product_images():
    """ Produktbilder importieren """

    for instance in Products.objects.all():

        product_images = ProductImages.objects.filter(product=instance)

        if len(product_images) == 0:

            for filename in os.listdir(os.path.join(settings.STATIC_DIR, 'app', 'img', 'product', instance.form)):

                if '1' in filename:
                    item = ProductImages.objects.create(product=instance, main_image=True)
                else:
                    item = ProductImages.objects.create(product=instance)

                directory = os.path.join(settings.STATIC_DIR, 'app', 'img', 'product', instance.form, filename)

                with open(directory, 'rb') as image_file:
                    item.img.save(os.path.basename(directory), File(image_file))

def import_effect_content(import_file, request):
    """ Cannabis Wirkungsweisen Inhalte importieren """

    workbook = load_workbook(import_file, data_only=True)

    ##### Natürliches Vorkommen #####
    worksheet = workbook.worksheets[0]

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        position = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Position', (row_offset - 2, 0))
        name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Bezeichnung', (row_offset - 2, 0))
        title = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Titel', (row_offset - 2, 0))
        teaser = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Teaser', (row_offset - 2, 0))
        teaser = teaser.replace("_x000D_", "") if teaser else ''
        content = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Inhalt als HTML', (row_offset - 2, 0))
        content = content.replace("_x000D_", "") if content else ''
        source = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Quelle', (row_offset - 2, 0))
        source = source.replace("_x000D_", "") if source else ''
        url_name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'URL Bezeichnung', (row_offset - 2, 0))
        color = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Farbschema', (row_offset - 2, 0))
        meta_title = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Meta:Titel', (row_offset - 2, 0))
        meta_description = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Meta:Beschreibung', (row_offset - 2, 0))
        meta_description = meta_description.replace("_x000D_", "") if meta_description else ''
        main_page = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Hauptseite', (row_offset - 2, 0))
        active = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Aktiv', (row_offset - 2, 0))

        CannabisEffects.objects.update_or_create(id=item_id, defaults={'position': int(position) if position else 0,
                                                                        'name': name if name else '',
                                                                        'title': title if title else '',
                                                                        'teaser': teaser if teaser else '',
                                                                        'content': content if content else '',
                                                                        'source': source if source else '',
                                                                        'url_name': url_name if url_name else '',
                                                                        'color': color if color else '',
                                                                        'meta_title': meta_title if meta_title else '',
                                                                        'meta_description': meta_description if meta_description else '',
                                                                        'main_page': True if main_page == '1' or main_page == 1 else False,
                                                                        'active': True if active == '1' or active == 1 else False
                                                                        })

        row_offset += 1

    # Create log entry
    log_entry = {
        'reference': 'Dataimport',
        'message': 'Wirkungsweisen-Datei importiert',
        'user': f'({ request.user.id }) { request.user.username }'
    }
    create_log(**log_entry)

def import_indications_content(import_file, request):
    """ Cannabis Indikationen Inhalte importieren """
    
    workbook = load_workbook(import_file, data_only=True)

    ##### Natürliches Vorkommen #####
    worksheet = workbook.worksheets[0]

    max_row = worksheet.max_row
    max_column = worksheet.max_column
    row_offset = 3

    while row_offset < max_row:

        item_id = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'ID', (row_offset - 2, 0))
        position = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Position', (row_offset - 2, 0))
        name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Bezeichnung', (row_offset - 2, 0))
        title = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Titel', (row_offset - 2, 0))
        matching_indication = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Passende Indikation', (row_offset - 2, 0))
        teaser = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Teaser', (row_offset - 2, 0))
        teaser = teaser.replace("_x000D_", "") if teaser else ''
        teaser_source = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Teaser Quelle', (row_offset - 2, 0))
        teaser_source = teaser_source.replace("_x000D_", "") if teaser_source else ''
        content = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Inhalt als HTML', (row_offset - 2, 0))
        content = content.replace("_x000D_", "") if content else ''
        source = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Quelle', (row_offset - 2, 0))
        source = source.replace("_x000D_", "") if source else ''
        url_name = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'URL Bezeichnung', (row_offset - 2, 0))
        meta_title = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Meta:Titel', (row_offset - 2, 0))
        meta_description = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Meta:Beschreibung', (row_offset - 2, 0))
        meta_description = meta_description.replace("_x000D_", "") if meta_description else ''
        active = ws_lookup_neighbour(worksheet, 3, max_column, max_row, 'Aktiv', (row_offset - 2, 0))

        try:
            indication = Indications.objects.get(name=matching_indication)
        except ObjectDoesNotExist:
            indication = None

        CannabisIndications.objects.update_or_create(id=item_id, defaults={'position': int(position) if position else 0,
                                                                        'name': name,
                                                                        'title': title,
                                                                        'teaser': teaser,
                                                                        'teaser_source': teaser_source,
                                                                        'content': content,
                                                                        'source': source,
                                                                        'url_name': url_name,
                                                                        'meta_title': meta_title,
                                                                        'meta_description': meta_description,
                                                                        'indication': indication,
                                                                        'active': True if active == '1' or active == 1 else False
                                                                        })

        row_offset += 1
    
    # Create log entry
    log_entry = {
        'reference': 'Dataimport',
        'message': 'Indikationen-Datei importiert',
        'user': f'({ request.user.id }) { request.user.username }'
    }
    create_log(**log_entry)

######### Export functions #########
def excel_date(date1):
    if date1.tzinfo is not None and date1.utcoffset() is not None:
        # Convert aware datetime to naive datetime
        date1 = date1.replace(tzinfo=None)
    temp = datetime(1899, 12, 30)
    delta = date1 - temp
    return float(delta.days) + (float(delta.seconds) / 86400)

def export_effect_content(ids):
    """ Export effect content to excel by ids """

    cannabis_effects = CannabisEffects.objects.filter(id__in=ids).order_by('position')

    # Create response
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachments; filename="Export_Wirkungsweisen_Content.xlsx"'

    # Create workbook
    workbook = Workbook()

    # Rename active sheet
    worksheet = workbook.active
    worksheet.title = 'Wirkungsweisen'

    # Styles
    header_font = styles.Font(color='FFFFFF', bold=True)
    header_fill = styles.PatternFill(fill_type='solid', fgColor='0066CC')
    font_fill = styles.PatternFill(fill_type='solid', fgColor='C5D9F1')
    font_border = styles.Border(left=styles.Side(border_style='thin', color='BFBFBF'),
                                right=styles.Side(border_style='thin', color='BFBFBF'),
                                top=styles.Side(border_style='thin', color='BFBFBF'),
                                bottom=styles.Side(border_style='thin', color='BFBFBF'),
                                )

    row_offset = 3

    cells = []

    # Headerline
    cells.append(worksheet.cell(row=row_offset, column=1, value='ID'))
    cells.append(worksheet.cell(row=row_offset, column=2, value='Bezeichnung'))
    cells.append(worksheet.cell(row=row_offset, column=3, value='Titel'))
    cells.append(worksheet.cell(row=row_offset, column=4, value='Position'))
    cells.append(worksheet.cell(row=row_offset, column=5, value='Teaser'))
    cells.append(worksheet.cell(row=row_offset, column=6, value='Inhalt als HTML'))
    cells.append(worksheet.cell(row=row_offset, column=7, value='Quelle'))
    cells.append(worksheet.cell(row=row_offset, column=8, value='URL Bezeichnung'))
    cells.append(worksheet.cell(row=row_offset, column=9, value='Meta:Titel'))
    cells.append(worksheet.cell(row=row_offset, column=10, value='Meta:Beschreibung'))
    cells.append(worksheet.cell(row=row_offset, column=11, value='Farbschema'))
    cells.append(worksheet.cell(row=row_offset, column=12, value='Hauptseite'))
    cells.append(worksheet.cell(row=row_offset, column=13, value='Aktiv'))

    # Change font and fill style of header
    for cell in cells:
        cell.font = header_font
        cell.fill = header_fill

    # Increase row_offset for content
    row_offset +=1

    # Wirte in workbook
    for index, item in enumerate(cannabis_effects):

        cells = []

        # Fill content
        cells.append(worksheet.cell(row=row_offset, column=1, value=item.id))
        cells.append(worksheet.cell(row=row_offset, column=2, value=item.name))
        cells.append(worksheet.cell(row=row_offset, column=3, value=item.title))
        cells.append(worksheet.cell(row=row_offset, column=4, value=item.position))
        cells.append(worksheet.cell(row=row_offset, column=5, value=item.teaser))
        cells.append(worksheet.cell(row=row_offset, column=6, value=item.content))
        cells.append(worksheet.cell(row=row_offset, column=7, value=item.source))
        cells.append(worksheet.cell(row=row_offset, column=8, value=item.url_name))
        cells.append(worksheet.cell(row=row_offset, column=9, value=item.meta_title))
        cells.append(worksheet.cell(row=row_offset, column=10, value=item.meta_description))
        cells.append(worksheet.cell(row=row_offset, column=11, value=item.color))
        cells.append(worksheet.cell(row=row_offset, column=12, value=1 if item.main_page else 0))
        cells.append(worksheet.cell(row=row_offset, column=13, value=1 if item.active else 0))

        if index % 2 == 0:
            for cell in cells:
                cell.fill = font_fill
                cell.border = font_border

        # Increase row_offset
        row_offset += 1

    # Save workbook
    workbook.save(response)

    return response

def export_indications_content(ids):
    """ Export indications content to excel by ids """

    cannabis_effects = CannabisIndications.objects.filter(id__in=ids).order_by('position')

    # Create response
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachments; filename="Export_Indikationen_Content.xlsx"'

    # Create workbook
    workbook = Workbook()

    # Rename active sheet
    worksheet = workbook.active
    worksheet.title = 'Indikationen'

    # Styles
    header_font = styles.Font(color='FFFFFF', bold=True)
    header_fill = styles.PatternFill(fill_type='solid', fgColor='0066CC')
    font_fill = styles.PatternFill(fill_type='solid', fgColor='C5D9F1')
    font_border = styles.Border(left=styles.Side(border_style='thin', color='BFBFBF'),
                                right=styles.Side(border_style='thin', color='BFBFBF'),
                                top=styles.Side(border_style='thin', color='BFBFBF'),
                                bottom=styles.Side(border_style='thin', color='BFBFBF'),
                                )

    row_offset = 3

    cells = []

    # Headerline
    cells.append(worksheet.cell(row=row_offset, column=1, value='ID'))
    cells.append(worksheet.cell(row=row_offset, column=2, value='Bezeichnung'))
    cells.append(worksheet.cell(row=row_offset, column=3, value='Titel'))
    cells.append(worksheet.cell(row=row_offset, column=4, value='Passende Indikation'))
    cells.append(worksheet.cell(row=row_offset, column=5, value='Position'))
    cells.append(worksheet.cell(row=row_offset, column=6, value='Teaser'))
    cells.append(worksheet.cell(row=row_offset, column=7, value='Teaser Quelle'))
    cells.append(worksheet.cell(row=row_offset, column=8, value='Inhalt als HTML'))
    cells.append(worksheet.cell(row=row_offset, column=9, value='Quelle'))
    cells.append(worksheet.cell(row=row_offset, column=10, value='URL Bezeichnung'))
    cells.append(worksheet.cell(row=row_offset, column=11, value='Meta:Titel'))
    cells.append(worksheet.cell(row=row_offset, column=12, value='Meta:Beschreibung'))
    cells.append(worksheet.cell(row=row_offset, column=13, value='Aktiv'))

    # Change font and fill style of header
    for cell in cells:
        cell.font = header_font
        cell.fill = header_fill

    # Increase row_offset for content
    row_offset +=1

    # Wirte in workbook
    for index, item in enumerate(cannabis_effects):

        cells = []

        # Fill content
        cells.append(worksheet.cell(row=row_offset, column=1, value=item.id))
        cells.append(worksheet.cell(row=row_offset, column=2, value=item.name))
        cells.append(worksheet.cell(row=row_offset, column=3, value=item.title))
        cells.append(worksheet.cell(row=row_offset, column=4, value=item.indication.name if item.indication else ''))
        cells.append(worksheet.cell(row=row_offset, column=5, value=item.position))
        cells.append(worksheet.cell(row=row_offset, column=6, value=item.teaser))
        cells.append(worksheet.cell(row=row_offset, column=7, value=item.teaser_source))
        cells.append(worksheet.cell(row=row_offset, column=8, value=item.content))
        cells.append(worksheet.cell(row=row_offset, column=9, value=item.source))
        cells.append(worksheet.cell(row=row_offset, column=10, value=item.url_name))
        cells.append(worksheet.cell(row=row_offset, column=11, value=item.meta_title))
        cells.append(worksheet.cell(row=row_offset, column=12, value=item.meta_description))
        cells.append(worksheet.cell(row=row_offset, column=13, value=1 if item.active else 0))

        if index % 2 == 0:
            for cell in cells:
                cell.fill = font_fill
                cell.border = font_border

        # Increase row_offset
        row_offset += 1

    # Save workbook
    workbook.save(response)

    return response

def export_order_products(ids):
    """ Export all order products """

    order_products = OrderProducts.objects.filter(id__in=ids, order__ordered=True).order_by('-order__order_time', 'order__number')

    # Create response
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachments; filename="Export_Bestellprodukte.xlsx"'

    # Create workbook
    workbook = Workbook()

    # Rename active sheet
    worksheet = workbook.active
    worksheet.title = 'Bestellprodukte'

    # Styles
    header_font = styles.Font(color='FFFFFF', bold=True)
    header_fill = styles.PatternFill(fill_type='solid', fgColor='0066CC')
    font_fill = styles.PatternFill(fill_type='solid', fgColor='C5D9F1')
    font_border = styles.Border(left=styles.Side(border_style='thin', color='BFBFBF'),
                                right=styles.Side(border_style='thin', color='BFBFBF'),
                                top=styles.Side(border_style='thin', color='BFBFBF'),
                                bottom=styles.Side(border_style='thin', color='BFBFBF'),
                                )

    row_offset = 1

    header_lines = [
        'Bestellnummer',
        'Datum',
        'Status',
        'Bezahlmethode',
        'Zahlungsstatus',
        'Artikelnummer',
        'Name',
        'Kultivar',
        'Hersteller',
        'Lieferant',
        'Herkunftsland',
        'Zerkleinert',
        'Menge',
        'Einkaufspreis (Netto)',
        'Einkaufspreis (Brutto)',
        'Einkaufspreis Summe (Netto)',
        'Einkaufspreis Summe (Brutto)',
        'Interne Lieferkosten (Netto)',
        'Interne Lieferkosten (Brutto)',
        'SUMME EK (Netto)',
        'SUMME EK (Brutto)',
        'Aufschlag auf EK Preis Netto',
        'Apotheke',
        'Selbstzahler VK Preis (Netto)',
        'Selbstzahler VK Preis (Brutto)',
        'VK Summe (Netto)',
        'VK Summe (Brutto)',
        'Versandkosten (Netto)',
        'Versandkosten (Brutto)',
        'SUMME VK (Netto)',
        'SUMME VK (Brutto)',
    ]

    cells = []

    # Headerline
    for index, item in enumerate(header_lines):
        cells.append(worksheet.cell(row=row_offset, column=index + 1, value=item))

    # Change font and fill style of header
    for cell in cells:
        cell.font = header_font
        cell.fill = header_fill

    # Increase row_offset for content
    row_offset +=1

    # Wirte in workbook
    for index, item in enumerate(order_products):

        related_order_products_amount = OrderProducts.objects.filter(order=item.order).count()

        cells = []

        #### Fill content ####
        # Bestellnummer
        cells.append(worksheet.cell(row=row_offset, column=1, value=int(item.order.number) if item.order.number else ''))
        # Datum
        cells.append(worksheet.cell(row=row_offset, column=2, value=excel_date(item.order.order_time)))
        # Status
        cells.append(worksheet.cell(row=row_offset, column=3, value=item.order.get_status_display()))
        # Bezahlmethode
        cells.append(worksheet.cell(row=row_offset, column=4, value=item.order.get_payment_type_display()))
        # Zahlungsstatus
        cells.append(worksheet.cell(row=row_offset, column=5, value=item.order.get_payment_status_display()))
        # Artikelnummer
        cells.append(worksheet.cell(row=row_offset, column=6, value=int(item.product.number)))
        # Name
        cells.append(worksheet.cell(row=row_offset, column=7, value=item.product.name))
        # Kultivar
        cells.append(worksheet.cell(row=row_offset, column=8, value=item.product.cultivar.name if item.product.cultivar else 'NICHT VORHANDEN'))
        # Hersteller
        cells.append(worksheet.cell(row=row_offset, column=9, value=item.product.manufacturer.name if item.product.manufacturer else 'NICHT VORHANDEN'))
        # Lieferant
        cells.append(worksheet.cell(row=row_offset, column=10, value=item.product.supplier.name if item.product.supplier else 'NICHT VORHANDEN'))
        # Herkunftsland
        cells.append(worksheet.cell(row=row_offset, column=11, value=item.product.country_of_origin.name if item.product.country_of_origin else 'NICHT VORHANDEN'))
        # Zerkleinert
        cells.append(worksheet.cell(row=row_offset, column=12, value='x' if item.prepared else ''))
        # Menge
        cells.append(worksheet.cell(row=row_offset, column=13, value=item.amount))
        # Einkaufspreis (Netto)
        cells.append(worksheet.cell(row=row_offset, column=14, value=item.purchase_price))
        # Einkaufspreis (Brutto)
        cells.append(worksheet.cell(row=row_offset, column=15, value=round((item.purchase_price * 1.19), 2)))
        # Einkaufspreis Summe (Netto)
        cells.append(worksheet.cell(row=row_offset, column=16, value=round(item.purchase_price * item.amount, 2)))
        # Einkaufspreis Summe (Brutto)
        cells.append(worksheet.cell(row=row_offset, column=17, value=round((item.purchase_price * 1.19) * item.amount, 2)))
        # Interne Lieferkosten (Netto)
        cells.append(worksheet.cell(row=row_offset, column=18, value=round((item.order.intern_delivery_costs/related_order_products_amount) / 1.19, 2)))
        # Interne Lieferkosten (Brutto)
        cells.append(worksheet.cell(row=row_offset, column=19, value=round(item.order.intern_delivery_costs/related_order_products_amount, 2)))
        # SUMME EK (Netto)
        cells.append(worksheet.cell(row=row_offset, column=20, value=round(item.purchase_price * item.amount, 2)))
        # SUMME EK (Brutto)
        cells.append(worksheet.cell(row=row_offset, column=21, value=round((item.purchase_price * 1.19) * item.amount, 2)))
        # Aufschlag auf EK Preis Netto
        cells.append(worksheet.cell(row=row_offset, column=22, value=item.price_surcharge))
        # Apotheke
        cells.append(worksheet.cell(row=row_offset, column=23, value=item.order.pharmacy.name if item.order.pharmacy else 'NICHT VORHANDEN'))
        # Selbstzahler VK Preis (Netto)
        cells.append(worksheet.cell(row=row_offset, column=24, value=item.price_net))
        # Selbstzahler VK Preis (Brutto)
        cells.append(worksheet.cell(row=row_offset, column=25, value=item.price))
        # VK Summe (Netto)
        cells.append(worksheet.cell(row=row_offset, column=26, value=round((item.price_net * item.amount), 2)))
        # VK Summe (Brutto)
        cells.append(worksheet.cell(row=row_offset, column=27, value=round((item.price * item.amount), 2)))
        # Versandkosten (Netto)
        cells.append(worksheet.cell(row=row_offset, column=28, value=round((item.order.delivery_costs/related_order_products_amount) / 1.19, 2)))
        # Versandkosten (Brutto)
        cells.append(worksheet.cell(row=row_offset, column=29, value=round(item.order.delivery_costs/related_order_products_amount, 2)))
        # SUMME VK (Netto)
        cells.append(worksheet.cell(row=row_offset, column=30, value=round((item.price_net * item.amount) + ((item.order.delivery_costs/related_order_products_amount) / 1.19), 2)))
        # SUMME VK (Brutto)
        cells.append(worksheet.cell(row=row_offset, column=31, value=round((item.price * item.amount) + (item.order.delivery_costs/related_order_products_amount), 2)))

        if index % 2 == 0:
            for cell in cells:
                cell.fill = font_fill
                cell.border = font_border

        # Change number format
        number_cells = [13, 14, 15, 16, 17, 18, 19, 20, 21, 23, 24, 25, 26, 27, 28, 29, 30]

        for number_cell in number_cells:
            cells[number_cell].number_format = '0.00'

        cells[0].number_format = '0'
        cells[1].number_format = '0'
        cells[5].number_format = '0'
        cells[12].number_format = '0'

        # Increase row_offset
        row_offset += 1

    # Save workbook
    workbook.save(response)

    return response

def export_orders(ids):
    """ Export all orders """

    orders = Orders.objects.filter(id__in=ids, ordered=True).order_by('-order_time', 'number')

    # Create response
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachments; filename="Export_Bestellprodukte.xlsx"'

    # Create workbook
    workbook = Workbook()

    # Rename active sheet
    worksheet = workbook.active
    worksheet.title = 'Bestellprodukte'

    # Styles
    header_font = styles.Font(color='FFFFFF', bold=True)
    header_fill = styles.PatternFill(fill_type='solid', fgColor='0066CC')
    font_fill = styles.PatternFill(fill_type='solid', fgColor='C5D9F1')
    font_border = styles.Border(left=styles.Side(border_style='thin', color='BFBFBF'),
                                right=styles.Side(border_style='thin', color='BFBFBF'),
                                top=styles.Side(border_style='thin', color='BFBFBF'),
                                bottom=styles.Side(border_style='thin', color='BFBFBF'),
                                )

    row_offset = 1

    cells = []

    # Headline
    cells.append(worksheet.cell(row=row_offset, column=1, value='Bestellnummer'))
    cells.append(worksheet.cell(row=row_offset, column=2, value='Datum'))
    cells.append(worksheet.cell(row=row_offset, column=3, value='Status'))
    cells.append(worksheet.cell(row=row_offset, column=4, value='Bezahlmethode'))
    cells.append(worksheet.cell(row=row_offset, column=5, value='Zahlungsstatus'))
    cells.append(worksheet.cell(row=row_offset, column=5, value='Liefermethode'))
    cells.append(worksheet.cell(row=row_offset, column=6, value='Patientenart'))
    cells.append(worksheet.cell(row=row_offset, column=7, value='Vorname'))
    cells.append(worksheet.cell(row=row_offset, column=8, value='Nachname'))
    cells.append(worksheet.cell(row=row_offset, column=9, value='Straße'))
    cells.append(worksheet.cell(row=row_offset, column=10, value='Hausnummer'))
    cells.append(worksheet.cell(row=row_offset, column=11, value='PLZ'))
    cells.append(worksheet.cell(row=row_offset, column=12, value='Ort'))
    cells.append(worksheet.cell(row=row_offset, column=13, value='Zwischensumme (Netto)'))
    cells.append(worksheet.cell(row=row_offset, column=14, value='Mwst. Betrag'))
    cells.append(worksheet.cell(row=row_offset, column=15, value='Zwischensumme (Brutto)'))
    cells.append(worksheet.cell(row=row_offset, column=16, value='Mahngebühr'))
    cells.append(worksheet.cell(row=row_offset, column=17, value='BTM Gebühr'))
    cells.append(worksheet.cell(row=row_offset, column=18, value='Versandkosten'))
    cells.append(worksheet.cell(row=row_offset, column=19, value='Bearbeitungsgebühr Bezahlmethode'))
    cells.append(worksheet.cell(row=row_offset, column=20, value='Verkaufspreis (Brutto)'))
    cells.append(worksheet.cell(row=row_offset, column=21, value='Interne Lieferkosten'))

    # Change font and fill style of header
    for cell in cells:
        cell.font = header_font
        cell.fill = header_fill

    # Increase row_offset for content
    row_offset +=1

    # Wirte in workbook
    for index, item in enumerate(orders):
            
        cells = []

        # Fill content
        cells.append(worksheet.cell(row=row_offset, column=1, value=int(item.number) if item.number else ''))
        cells.append(worksheet.cell(row=row_offset, column=2, value=excel_date(item.order_time)))
        cells.append(worksheet.cell(row=row_offset, column=3, value=item.get_status_display()))
        cells.append(worksheet.cell(row=row_offset, column=4, value=item.get_payment_type_display()))
        cells.append(worksheet.cell(row=row_offset, column=5, value=item.get_payment_status_display()))
        cells.append(worksheet.cell(row=row_offset, column=6, value=item.get_delivery_type_display()))
        cells.append(worksheet.cell(row=row_offset, column=7, value=item.first_name))
        cells.append(worksheet.cell(row=row_offset, column=8, value=item.last_name))
        cells.append(worksheet.cell(row=row_offset, column=9, value=item.street))
        cells.append(worksheet.cell(row=row_offset, column=10, value=item.street_number))
        cells.append(worksheet.cell(row=row_offset, column=11, value=item.postalcode))
        cells.append(worksheet.cell(row=row_offset, column=12, value=item.city))
        cells.append(worksheet.cell(row=row_offset, column=13, value=item.subtotal))
        cells.append(worksheet.cell(row=row_offset, column=14, value=item.tax_amount))
        cells.append(worksheet.cell(row=row_offset, column=15, value=item.subtotal_brutto))
        cells.append(worksheet.cell(row=row_offset, column=16, value=item.reminder_fee))
        cells.append(worksheet.cell(row=row_offset, column=17, value=item.btm_fee))
        cells.append(worksheet.cell(row=row_offset, column=18, value=item.delivery_costs))
        cells.append(worksheet.cell(row=row_offset, column=19, value=item.payment_fee))
        cells.append(worksheet.cell(row=row_offset, column=20, value=item.total))
        cells.append(worksheet.cell(row=row_offset, column=21, value=item.intern_delivery_costs))

        if index % 2 == 0:
            for cell in cells:
                cell.fill = font_fill
                cell.border = font_border

        cells[0].number_format = '0'
        cells[1].number_format = '0'
        cells[10].number_format = '0'
        cells[12].number_format = '0.00'
        cells[13].number_format = '0.00'
        cells[14].number_format = '0.00'
        cells[15].number_format = '0.00'
        cells[16].number_format = '0.00'
        cells[17].number_format = '0.00'
        cells[18].number_format = '0.00'
        cells[19].number_format = '0.00'
        cells[20].number_format = '0.00'

        # Increase row_offset
        row_offset += 1

    # Save workbook
    workbook.save(response)

    return response

######### User Functions #########
def count_notifications(user):
    """ Count total notifications """

    amount = 0

    amount += count_product_request_notifications(user)

    return amount

def count_product_request_notifications(user):
    """ Count product requests, which are approved """

    try:
        customer = Customers.objects.get(user=user)
        return ProductRequest.objects.filter(customer=customer, status='approved', ordered=False, declined=False).count()
    except:
        return 0

######### Cart #########
def add_product_to_cart(order, product_id, amount, prepared=False, user_is_authenticated=False, from_product_request=False, requested_product=None):
    """ Add product to cart by product.id """

    data = {}

    try:
        product = Products.objects.get(id=product_id)
        product_price = ProductPrices.objects.get(product=product, pharmacy=order.pharmacy)

        # if from_product_request:

        order_product, created = OrderProducts.objects.get_or_create(
                                    order=order,
                                    product=product,
                                    prepared=prepared,
                                    is_requested=from_product_request,
                                    requested_product=requested_product
                                )
        
        if not from_product_request:

            if product.form == 'flower' and amount < 5:
                amount = 5

            order_product.amount += amount
        else:
            order_product.amount = amount

        if order_product.amount > 100:
            order_product.amount = 100

        order_product.save()

        data = {
            'object_not_exist': False,
            'avaliable': True,
            'created': created,
            'id': order_product.id,
            'productId': product.id,
            'name': order_product.product.name,
            'amount': order_product.amount,
            'form': order_product.product.get_form_display(),
            'supplier': order_product.product.supplier.name,
            'prepared': prepared,
            'preparedVisible': True if product.form == 'flower' else False,
            'unit': 'g' if order_product.product.form == 'flower' else 'Einheiten',
            'img': ProductImages.objects.filter(product=product, main_image=True).first().img.url if ProductImages.objects.filter(product=product, main_image=True) else '',
            'total': custom_currency_format(order_product.total) if user_is_authenticated else 0,
            'total_value': order_product.total if user_is_authenticated else 0,
        }

        # else:

        #     data = {
        #         'id': product.id,
        #         'name': product.name,
        #         'amount': amount,
        #         'prepared': prepared,
        #         'preparedVisible': True if product.form == 'flower' else False,
        #         'img': ProductImages.objects.filter(product=product, main_image=True).first().img.url if ProductImages.objects.filter(product=product, main_image=True) else '',
        #         'status': product_price.status,
        #         'avaliable': False,
        #     }

    except Products.DoesNotExist:

        data['object_not_exist'] = True

    return data

def get_order_details(order_id):
    """ Get all Order details with prices """

    order = Orders.objects.get(id=order_id)
    order_products = OrderProducts.objects.filter(order=order)

    products = []

    for order_product in order_products:
        products.append({
            'id': order_product.id,
            'name': order_product.product.name,
            'form': order_product.product.form,
            'amount': order_product.amount,
            'price': custom_currency_format(order_product.price),
            'total': custom_currency_format(order_product.total),
            'prepared': 'Ver#ndert' if order_product.prepared else 'Unverändert',
        })

    data = {
        'id': order.id,
        'first_name': order.first_name,
        'last_name': order.last_name,
        'street': order.street,
        'street_number': order.street_number,
        'postalcode': order.postalcode,
        'city': order.city,
        'phone_number': order.phone_number,
        'email': order.email_address,
        'country': order.country,
        'subtotal': custom_currency_format(order.subtotal),
        'subtotal_brutto': custom_currency_format(order.subtotal_brutto),
        'tax_amount': custom_currency_format(order.tax_amount),
        'discount': custom_currency_format(order.discount),
        'delivery_costs_label': 'Versandkosten' if order.delivery_type else 'Zzgl. Versandkosten',
        'delivery_costs': custom_currency_format(order.delivery_costs) if order.delivery_costs else '-- €',
        'total': custom_currency_format(order.total),
        'insurance_participation': custom_currency_format(order.insurance_participation),
        'co_payment': custom_currency_format(order.co_payment),
        'amount_payable': custom_currency_format(order.amount_payable),
        'products': products,
        'date': order.order_time.strftime('%d.%m.%Y'),
        'payment_status': order.payment_status,
        'payment_status_display': order.get_payment_status_display(),
        'recipe_status': order.recipe_status,
        'recipe_status_display': order.get_recipe_status_display(),
        'status': order.status,
        'status_display': order.get_status_display(),
    }

    return data

def get_product_details(product_id, user_is_authenticated=False):
    """ Get order product details by id """
 
    try:
        order_product = OrderProducts.objects.get(id=product_id)

        data = {
            'id': order_product.id,
            'name': order_product.product.name,
            'amount': order_product.amount,
            'prepared': order_product.prepared,
            'unit': 'g' if order_product.product.form == 'flower' else 'ml',
            'total': custom_currency_format(order_product.total) if user_is_authenticated else 0,
            'total_value': order_product.total if user_is_authenticated else 0,
        }
    except ObjectDoesNotExist:
        data = None

    return data

######### Products #########
def filter_products(parameters):
    """ Filter product by parameters """

    products = Products.objects.filter(active=True)

    products = products.annotate(
        custom_status=Case(
            When(status='0', then=Value(1)),
            default=Value(0),
            output_field=IntegerField(),
        )
    )

    # Verfügbarkeit
    if 'status' in parameters:
        filter_status = parameters.get('status').split(',')
        products = products.filter(status__in=get_values_by_display_nams(ProductStatusChoices, filter_status))

    # Genetik
    if 'genetic' in parameters:
        filter_genetics = parameters.get('genetic').split(',')
        products = products.filter(genetics__name__in=filter_genetics)

    # Terpene
    if 'terpene' in parameters:
        filter_terpene = parameters.get('terpene').split(',')
        products = products.filter(main_terpene__name__in=filter_terpene)

    # Terpenenwirkungen
    if 'terpene_effects' in parameters:
        filter_effects = parameters.get('terpene_effects').split(',')
        terpene_effects_ids = list(TerpeneEffects.objects.filter(name__in=filter_effects).values_list('id', flat=True))
        products = products.filter(main_terpene__terpene_effect__id__in=terpene_effects_ids)

    # Indikationen
    if 'indications' in parameters:
        filter_indications = parameters.get('indications').split(',')
        indications_ids = list(Indications.objects.filter(name__in=filter_indications).values_list('id', flat=True))
        products = products.filter(main_terpene__indications__id__in=indications_ids)

    # Hersteller
    if 'manufacturer' in parameters:
        filter_manufacturer = parameters.get('manufacturer').split(',')
        manufacturer_ids = list(Manufacturer.objects.filter(name__in=filter_manufacturer).values_list('id', flat=True))
        products = products.filter(manufacturer__id__in=manufacturer_ids)

    # Kultivar
    if 'cultivar' in parameters:
        filter_cultivar = parameters.get('cultivar').split(',')
        cultivar_ids = list(Cultivar.objects.filter(name__in=filter_cultivar).values_list('id', flat=True))
        products = products.filter(cultivar__id__in=cultivar_ids)

    # Form
    if 'form' in parameters:
        filter_forms = parameters.get('form').split(',')
        products = products.filter(form__in=get_values_by_display_nams(CannabisFormChoices, filter_forms))

    # Sortieren
    if 'sort' in parameters:
        sort_type = parameters.get('sort')

        if sort_type == 'Relevanteste':
            products = products.order_by('-special_offer', '-form', '-status', '-priority')

        if sort_type == 'Name aufsteigend':
            products = products.order_by('custom_status', 'name')

        if sort_type == 'Name absteigend':
            products = products.order_by('custom_status', '-name')

        if sort_type == 'Niedrigster Preis':
            products = products.order_by('custom_status', 'self_payer_selling_price_brutto')

        if sort_type == 'Höchster Preis':
            products = products.order_by('custom_status', '-self_payer_selling_price_brutto')

        if sort_type == 'THC aufsteigend':
            products = products.order_by('custom_status', 'thc_value')

        if sort_type == 'THC absteigend':
            products = products.order_by('custom_status', '-thc_value')

        if sort_type == 'CBD aufsteigend':
            products = products.order_by('custom_status', 'max_cbd_value')

        if sort_type == 'CBD absteigend':
            products = products.order_by('custom_status', '-max_cbd_value')

    else:
        products = products.order_by('-special_offer', '-status', '-priority')

    # Suchwort
    if 'search' in parameters:
        search_word = parameters.get('search')

        values = search_word.split(' ')
        values = [value for value in values if value]

        q_objects = Q()

        for value in values:

            q_objects &= (
                Q(name__icontains=value) |
                Q(cultivar__name__icontains=value) |
                Q(genetics__name__icontains=value) |
                Q(manufacturer__name__icontains=value) |
                Q(main_terpene__name__icontains=value) |
                Q(main_terpene__terpene_effect__name__icontains=value)
            )

        products = products.filter(q_objects).distinct()

    return products.distinct()

def get_filters(products):
    """ Get all filters """

    filters = {}

    terpene_ids = products.values_list('main_terpene__id', flat=True).distinct()
    terpene = Terpene.objects.filter(id__in=terpene_ids).order_by('name')

    terpene_effect_ids = terpene.values_list('terpene_effect__id', flat=True).distinct()
    terpene_effects = TerpeneEffects.objects.filter(id__in=terpene_effect_ids).order_by('name')

    indication_ids = terpene.values_list('indications__id', flat=True).distinct()
    indications = Indications.objects.filter(id__in=indication_ids).order_by('name')

    cultivas_ids = products.values_list('cultivar__id', flat=True).distinct()
    cultivas = Cultivar.objects.filter(id__in=cultivas_ids).order_by('name')

    manufacturer_ids = products.values_list('manufacturer__id', flat=True).distinct()
    manufacturers = Manufacturer.objects.filter(id__in=manufacturer_ids).order_by('name')

    genetic_ids = products.values_list('genetics__id', flat=True).distinct()
    genetics = Genetics.objects.filter(id__in=genetic_ids).order_by('name')

    filters['cultivas'] = cultivas
    filters['manufacturers'] = manufacturers
    filters['genetics'] = genetics
    filters['terpene'] = terpene
    filters['terpene_effects'] = terpene_effects
    filters['indications'] = indications

    return filters

def check_min_order_amount(order):
    """ Check if order has minimum order amount """

    price_settings = PriceSettings.objects.first()

    extract_products = OrderProducts.objects.filter(order=order, product__form='extract')

    if extract_products:
        return True

    return order.total_product_amount >= price_settings.min_order_amount

def sum_product_stock_ordered_amount(stock_product, pharmacy):
    """ Get amount of ordered product """

    amount = ProductOrders.objects.filter(product=stock_product.product, pharmacy=pharmacy).aggregate(total=Sum('amount'))['total']

    return amount if amount else 0

def sum_product_ordered_amount(product, pharmacy):
    """ Get amount of ordered product """

    amount = OrderProducts.objects.filter(
        product=product, order__pharmacy=pharmacy,
        order__status__in=['ordered', 'process', 'process_waiting', 'queries', 'clarified'],
        order__ordered=True
        ).aggregate(total=Sum('amount'))['total']

    return amount if amount else 0

######### Order functions #########
def check_status_for_mail(order, status_name, old_status, new_status, request):
    """ Check order status by name for mailing """

    if old_status != new_status:

        # New Payment status
        if status_name == 'payment_status':

            if new_status == 'received':
                send_payment_status_received(order, request)
                try:
                    invoice = Invoices.objects.get(order=order, cancellation_invoice=False, canceled=False, order_invoice=True)
                    send_invoice_to_customer(invoice.id, request)
                except ObjectDoesNotExist:
                    create_log(
                        reference=f'check_status_for_mail - Order-Nr. { order.number } - Send invoice to customer if new payment status is received',
                        message='No invoice found',
                        category='error',
                        user=f'({ request.user.id }) { request.user.username }'
                    )

        # New Online Recipe status
        if status_name == 'online_recipe_status':

            if new_status == 'checked' and order.payment_type in ('prepayment', 'payment_by_invoice'):
                try:
                    invoice = Invoices.objects.get(order=order, cancellation_invoice=False, canceled=False).id
                    send_invoice_to_customer(invoice, request)
                except invoice.DoesNotExist:
                    pass

            if new_status == 'incorrect':
                send_recipe_status_incorrect(order)

        # New Recipe status
        if status_name == 'recipe_status':

            if new_status == 'received':
                send_recipe_status_received(order)

            if new_status == 'incorrect':
                send_recipe_status_incorrect(order)

        # New Order status
        if status_name == 'status':

            if new_status == 'ready_for_pickup':
                send_order_ready_for_pickup(order, request)

            if new_status == 'shipped':
                send_order_status_shipped(order)

            if new_status == 'cancelled':
                send_order_status_cancelled(order)

    return True

def create_new_invoice(order):
    """ Create new invoice for order """

    invoice = None
    
    try:
        invoice = Invoices.objects.get(order=order, cancellation_invoice=False, canceled=False, order_invoice=True)
        invoice.canceled = True
        invoice.save()

        create_cancelation_invoice(invoice)
    except ObjectDoesNotExist:
        pass

    new_invoice = Invoices.objects.create(order=order, order_invoice=True)

    if invoice:
        new_invoice.pro_forma_invoice = invoice.pro_forma_invoice
        new_invoice.save()

    return new_invoice

def create_cancelation_invoice(invoice):

    inovice_items = InvoiceItems.objects.filter(invoice=invoice)

    cancelation_invoice = Invoices.objects.create(
        cancellation_invoice    = True,
        pro_forma_invoice       = invoice.pro_forma_invoice,
        own_company_name        = invoice.own_company_name,
        own_street              = invoice.own_street,
        own_street_number       = invoice.own_street_number,
        own_postalcode          = invoice.own_postalcode,
        own_city                = invoice.own_city,
        customer                = invoice.customer,
        recipe_number           = invoice.recipe_number,
        order                   = invoice.order,
        customer_type           = invoice.order.customer_type,
        payment_type            = invoice.order.payment_type,
        customer_number         = invoice.customer.id,
        first_name              = invoice.order.first_name,
        last_name               = invoice.order.last_name,
        street                  = invoice.order.street,
        street_number           = invoice.order.street_number,
        postalcode              = invoice.order.postalcode,
        city                    = invoice.order.city,
        tax_rate                = invoice.tax_rate,
        delivery_costs          = invoice.delivery_costs,
        discount                = invoice.discount,
        subtotal                = invoice.subtotal,
        tax_amount              = invoice.tax_amount,
        total                   = invoice.total,
        co_payment              = invoice.co_payment,
        insurance_participation = invoice.insurance_participation,
        amount_payable          = invoice.amount_payable,
        order_invoice           = invoice.order_invoice,
    )

    for inovice_item in inovice_items:
        InvoiceItems.objects.create(
            invoice=cancelation_invoice,
            product_number=inovice_item.product_number,
            product=inovice_item.product,
            form=inovice_item.form,
            amount=inovice_item.amount,
            price=inovice_item.price,
            discount=inovice_item.discount,
            total=inovice_item.total
        )

def update_order_prices(order):
    """ Updates all order prices and existing invoices based on model rules """

    order.save()

    try:
        invoice = Invoices.objects.get(order=order, cancellation_invoice=False, canceled=False)
        invoice.delivery_costs = order.delivery_costs
        invoice.discount = order.discount
        invoice.subtotal = order.subtotal
        invoice.tax_amount = order.tax_amount
        invoice.total = order.total
        invoice.reminder_fee = order.reminder_fee
        invoice.co_payment = order.co_payment
        invoice.insurance_participation = order.insurance_participation
        invoice.amount_payable = order.amount_payable
        invoice.save()
    except ObjectDoesNotExist:
        pass

def check_for_payment_reminder(order):
    """ Check if order need payment reminder """

    if order.status in ['delivered']\
        and order.payment_status in ['pending']\
        and not order.invoice_reminder_send_on:

        _days_ago = timezone.now().date() - timedelta(days=15)

        if order.delivered_on.date() <= _days_ago:
            send_payment_reminder(order)

    elif order.status in ['delivered']\
        and order.payment_status in ['invoice_reminder']\
        and not order.last_reminder_send_on:

        _days_ago = timezone.now().date() - timedelta(days=6)

        if order.invoice_reminder_send_on and order.invoice_reminder_send_on.date() <= _days_ago:
            send_last_payment_reminder(order)

    elif order.status in ['delivered']\
        and order.payment_status in ['last_reminder']:

        _days_ago = timezone.now().date() - timedelta(days=6)

        if order.invoice_reminder_send_on and order.last_reminder_send_on.date() <= _days_ago:
            send_overdue_mail(order)

def block_customer(customer, till, reaseon):
    """ Block customer """

    customer.blocked = True
    customer.blocked_date = till
    customer.blocked_reason = reaseon
    customer.save()

def confirm_created_order(order_id, request):
    """ Confirm created order """

    order = Orders.objects.get(id=order_id)
    order.status = 'ordered'
    order.ordered = True
    order.save()

    try:
        send_order_confirmation(order.id, request)
    except Exception as e:
        create_log(
            reference='user_order_functions_v1- saveOrder',
            message=f'Order confirmation email not sent',
            user=f'({ request.user.id }) { request.user.username }',
            category='error',
            stack_trace=str(e)
        )

    try:
        invoice = Invoices.objects.get(order=order, cancellation_invoice=False, canceled=False)
        send_invoice_to_customer(invoice.id, request)
    except Exception as e:
        create_log(
            reference='user_order_functions_v1- saveOrder',
            message=f'Invoice not found in order confirm',
            user=f'({ request.user.id }) { request.user.username }',
            category='error',
            stack_trace=str(e)
        )

######### Dashboard functions #########
def dashboard_filter_orders(parameters, pharmacy):
    
    """ Filter order by parameters """

    orders = Orders.objects.filter(Q(ordered=True) | Q(status='started')).order_by('-order_time')

    if 'status' in parameters:

        selected_statuses = []

        for item in OrderStatusChoices:
            if item[1] in parameters.get('status').split(','):
                selected_statuses.append(item[0])

        orders = orders.filter(status__in=selected_statuses)

    if 'payment_status' in parameters:

        selected_statuses = []

        for item in PaymentStatusChoices:
            if item[1] in parameters.get('payment_status').split(','):
                selected_statuses.append(item[0])

        orders = orders.filter(payment_status__in=selected_statuses)

    if 'online_recipe_status' in parameters:
            
            selected_statuses = []
    
            for item in OnlineRecipeStatusChoices:
                if item[1] in parameters.get('online_recipe_status').split(','):
                    selected_statuses.append(item[0])
    
            orders = orders.filter(online_recipe_status__in=selected_statuses)

    if 'recipe_status' in parameters:
            
        selected_statuses = []

        for item in RecipeStatusChoices:
            if item[1] in parameters.get('recipe_status').split(','):
                selected_statuses.append(item[0])

        orders = orders.filter(recipe_status__in=selected_statuses)

    if 'payment_type' in parameters:
            
        selected_statuses = []

        for item in PaymentTypeChoices:
            if item[1] in parameters.get('payment_type').split(','):
                selected_statuses.append(item[0])

        orders = orders.filter(payment_type__in=selected_statuses)

    if 'customer_type' in parameters:

        selected_statuses = []

        for item in CustomerTypeChoices:
            if item[1] in parameters.get('customer_type').split(','):
                selected_statuses.append(item[0])

        orders = orders.filter(customer_type__in=selected_statuses)

    if 'delivery_type' in parameters:
            
        selected_statuses = []

        for item in DeliveryTypeChoices:
            if item[1] in parameters.get('delivery_type').split(','):
                selected_statuses.append(item[0])

        orders = orders.filter(delivery_type__in=selected_statuses)

    if 'search_product' in parameters:

        product_name = parameters.get('search_product')

        product = Products.objects.filter(name=product_name).first()
        order_products = OrderProducts.objects.filter(product=product)
        order_ids = order_products.values_list('order', flat=True)

        orders = orders.filter(id__in=order_ids).distinct()

    if 'search_invoice' in parameters:
            
        invoice_number = parameters.get('search_invoice')

        invoices = Invoices.objects.filter(invoice_number__icontains=invoice_number)

        order_ids = invoices.values_list('order', flat=True)

        orders = orders.filter(id__in=order_ids).distinct()

    if 'search' in parameters:

        search_word = parameters.get('search')

        values = search_word.split(' ')
        values = [value for value in values if value]

        q_objects = Q()

        for value in values:

            q_objects &= (
                Q(customer__user__first_name__icontains=value) |
                Q(customer__user__last_name__icontains=value) |
                Q(customer__user__email__icontains=value) |
                Q(number__icontains=value)
            )

        orders = orders.filter(q_objects).distinct()

    return orders.filter(pharmacy=pharmacy).distinct()

def dashboard_filter_prescription_orders(parameters, pharmacy):
    """ Filter prescription orders by parameters """

    orders = Orders.objects.filter(status='in_review').order_by('-id')

    if 'search' in parameters:
            
        search_word = parameters.get('search')

        values = search_word.split(' ')
        values = [value for value in values if value]

        q_objects = Q()

        for value in values:

            q_objects &= (
                Q(customer__user__first_name__icontains=value) |
                Q(customer__user__last_name__icontains=value) |
                Q(customer__user__email__icontains=value) |
                Q(number__icontains=value)
            )

        orders = orders.filter(q_objects).distinct()

    return orders.filter(pharmacy=pharmacy).distinct()

def dashboard_filter_order_products(parameters, pharmacy):
    """ Filter ordered products """

    order_products = OrderProducts.objects.filter(order__ordered=True).order_by('-order__order_time')

    if 'status' in parameters:

        selected_statuses = []

        for item in OrderStatusChoices:
            if item[1] in parameters.get('status').split(','):
                selected_statuses.append(item[0])

        order_products = order_products.filter(order__status__in=selected_statuses)

    if 'manufacturer' in parameters:

        filter_manufacturer = parameters.get('manufacturer').split(',')
        manufacturer_ids = list(Manufacturer.objects.filter(name__in=filter_manufacturer).values_list('id', flat=True))
        order_products = order_products.filter(product__manufacturer__id__in=manufacturer_ids)

    if 'supplier' in parameters:

        filter_supplier = parameters.get('supplier').split(',')
        supplier_ids = list(Supplier.objects.filter(name__in=filter_supplier).values_list('id', flat=True))
        order_products = order_products.filter(product__supplier__id__in=supplier_ids)

    if 'search' in parameters:

        search_word = parameters.get('search')

        values = search_word.split(' ')
        values = [value for value in values if value]

        q_objects = Q()

        for value in values:

            q_objects &= (
                Q(product__number__icontains=value) |
                Q(product__name__icontains=value) |
                Q(product__manufacturer__name__icontains=value) |
                Q(product__supplier__name__icontains=value) |
                Q(order__number__icontains=value) |
                Q(order__customer__user__first_name__icontains=value) |
                Q(order__customer__user__last_name__icontains=value) |
                Q(order__customer__user__email__icontains=value)
            )

        order_products = order_products.filter(q_objects).distinct()

    return order_products.filter(order__pharmacy=pharmacy).distinct()

def dashboard_filter_customers(parameters, pharmacy):
    
    """ Filter product by parameters """

    customers = Customers.objects.filter(pharmacies=pharmacy).order_by('-id')

    if 'search' in parameters:

        search_word = parameters.get('search')

        values = search_word.split(' ')
        values = [value for value in values if value]

        q_objects = Q()

        for value in values:

            q_objects &= (
                Q(user__first_name__icontains=value) |
                Q(user__last_name__icontains=value) |
                Q(user__email__icontains=value)
            )


        customers = customers.filter(q_objects).distinct()

    return customers.distinct()

def dashboard_filter_products(parameters):
    """ Filter product by parameters """

    products = Products.objects.all().order_by('-id')

    if 'status' in parameters:

        selected_statuses = []

        for item in ProductStatusChoices:
            if item[1] in parameters.get('status').split(','):
                selected_statuses.append(item[0])

        products = products.filter(status__in=selected_statuses)

    if 'manufacturer' in parameters:

        filter_manufacturer = parameters.get('manufacturer').split(',')
        manufacturer_ids = list(Manufacturer.objects.filter(name__in=filter_manufacturer).values_list('id', flat=True))
        products = products.filter(manufacturer__id__in=manufacturer_ids)

    if 'supplier' in parameters:

        filter_supplier = parameters.get('supplier').split(',')
        supplier_ids = list(Supplier.objects.filter(name__in=filter_supplier).values_list('id', flat=True))
        products = products.filter(supplier__id__in=supplier_ids)
    
    if 'cultivar' in parameters:

        filter_cultivar = parameters.get('cultivar').split(',')
        cultivar_ids = list(Cultivar.objects.filter(name__in=filter_cultivar).values_list('id', flat=True))
        products = products.filter(cultivar__id__in=cultivar_ids)

    if 'search' in parameters:

        search_word = parameters.get('search')

        values = search_word.split(' ')
        values = [value for value in values if value]

        q_objects = Q()

        for value in values:

            q_objects &= (
                Q(number__icontains=value) |
                Q(name__icontains=value) |
                Q(manufacturer__name__icontains=value) |
                Q(supplier__name__icontains=value)
            )

        products = products.filter(q_objects).distinct()

    return products.distinct()

def dashboard_filter_stock_products(parameters, pharmacy):
    """ Filter stock products by parameters"""

    products = StockProducts.objects.filter(pharmacy=pharmacy).order_by('product__name')

    if 'status' in parameters:
            
        selected_statuses = []

        for item in StockStatusChoices:
            if item[1] in parameters.get('status').split(','):
                selected_statuses.append(item[0])

        products = products.filter(status__in=selected_statuses)

    if 'manufacturer' in parameters:
            
        filter_manufacturer = parameters.get('manufacturer').split(',')
        manufacturer_ids = list(Manufacturer.objects.filter(name__in=filter_manufacturer).values_list('id', flat=True))
        products = products.filter(product__manufacturer__id__in=manufacturer_ids)

    if 'supplier' in parameters:

        filter_supplier = parameters.get('supplier').split(',')
        supplier_ids = list(Supplier.objects.filter(name__in=filter_supplier).values_list('id', flat=True))
        products = products.filter(product__supplier__id__in=supplier_ids)

    if 'search' in parameters:

        search_word = parameters.get('search')

        values = search_word.split(' ')
        values = [value for value in values if value]

        q_objects = Q()

        for value in values:

            q_objects &= (
                Q(product__number__icontains=value) |
                Q(product__name__icontains=value) |
                Q(product__manufacturer__name__icontains=value) |
                Q(product__supplier__name__icontains=value)
            )

        products = products.filter(q_objects).distinct()
    
    # Berechne die Gesamtmenge für jedes Produkt und erhalte die relevanten Felder
    products_with_total_amount = products.values(
        'product__id', 'product__name',
    ).annotate(total_amount=Sum('amount')).order_by('product__name')

    # Erstelle ein Dictionary, das die Gesamtmenge für jedes Produkt speichert
    product_total_amount_dict = {product['product__id']: product['total_amount'] for product in products_with_total_amount}

    # Alle StockProducts laden und nach Produkt-ID gruppieren
    all_livestock_products = products.select_related('product').order_by('product__name')

    # Erstelle eine Liste der eindeutigen Produkte mit den berechneten Gesamtmengen
    seen_products = set()
    products = []
    for livestock_product in all_livestock_products:
        product_id = livestock_product.product.id
        if product_id not in seen_products:
            seen_products.add(product_id)

            try:
                threshold = ProductThresholds.objects.get(product_id=product_id, pharmacy=pharmacy)
            except ObjectDoesNotExist:
                threshold = None

            # Calculate the total ordered amount
            ordered_amount = sum_product_ordered_amount(livestock_product.product, livestock_product.pharmacy)

            # Calculate enough_amount
            enough_amount = (
                (ordered_amount <= product_total_amount_dict[product_id]) and
                (threshold.threshold <= product_total_amount_dict[product_id] if threshold else True)
            )

            # Get highest status of stock_products
            highest_status_product = StockProducts.objects.filter(product=livestock_product.product, pharmacy=livestock_product.pharmacy).order_by('-status').first()

            product_info = {
                'id': livestock_product.id,
                'product': livestock_product.product,
                'total_amount': product_total_amount_dict[product_id],
                'pharmacy': livestock_product.pharmacy,
                'status': highest_status_product.status,
                'status_display': highest_status_product.get_status_display(),
                'amount_status': livestock_product.amount_status,
                'threshold': threshold.threshold if threshold else 'null',
                'enough_amount': enough_amount,
                'ordered_amount': ordered_amount,
            }
            products.append(product_info)

    products = sorted(products, key=lambda x: x['enough_amount'])

    return products

def dashboard_filter_stock_packages(parameters, pharmacy):
    """ Filter stock packages by parameters"""
    
    packages = Packages.objects.filter(pharmacy=pharmacy).order_by('id')

    # Filtern der StockPackages für die TEST APOTHEKE
    stock_packages = Packages.objects.filter(pharmacy=pharmacy)

    # Gruppieren und Annotieren
    grouped_stock_packages = stock_packages.values(
        'name', 'manufacturer', 'size'
    ).annotate(
        total_amount=Sum('amount')
    )

    # Erstellen der endgültigen Datenstruktur
    packages = []
    for group in grouped_stock_packages:
        name = group['name']
        manufacturer_id = group['manufacturer']
        size = group['size']

        # Finde alle Pakete in der aktuellen Gruppe
        packages_in_group = stock_packages.filter(
            name=name,
            manufacturer=manufacturer_id,
            size=size
        ).values('id', 'batch_number', 'amount', 'created_at')

        manufacturer = PackageManufacturers.objects.get(id=manufacturer_id)

        packages.append({
            'name': name,
            'manufacturer': manufacturer,
            'size': size,
            'total_amount': group['total_amount'],
            'packages': list(packages_in_group)
        })

    if 'manufacturer' in parameters:
            
        filter_manufacturer = parameters.get('manufacturer').split(',')
        manufacturer_ids = list(PackageManufacturers.objects.filter(name__in=filter_manufacturer).values_list('id', flat=True))
        packages = packages.filter(package__manufacturer__id__in=manufacturer_ids)

    if 'size' in parameters:

        filter_size = parameters.get('size').split(',')
        packages = packages.filter(size=filter_size)
        
    if 'search' in parameters:

        search_word = parameters.get('search')

        values = search_word.split(' ')
        values = [value for value in values if value]

        q_objects = Q()

        for value in values:

            q_objects &= (
                Q(package__number__icontains=value) 
            )

        packages = packages.filter(q_objects).distinct()

    return packages

def dashboard_filter_email_recipients(parameters, pharmacy):
    """ Filter email recipients by parameters """

    recipients = EmailRecipients.objects.all().order_by('category')

    if 'category' in parameters:
            
        selected_categories = []

        for item in EmailRecipientCategories:
            if item[1] in parameters.get('category').split(','):
                selected_categories.append(item[0])

        recipients = recipients.filter(category__in=selected_categories)


    if 'search' in parameters:

        search_word = parameters.get('search')

        values = search_word.split(' ')
        values = [value for value in values if value]

        q_objects = Q()

        for value in values:

            q_objects &= (
                Q(pharmacy__name__icontains=value) |
                Q(name__icontains=value) |
                Q(email__icontains=value)
            )

        recipients = recipients.filter(q_objects).distinct()

    return recipients.filter(pharmacy=pharmacy).distinct()

def dashboard_filter_invoices(parameters, pharmacy):
    """ Filter invoices by parameters """

    invoices = Invoices.objects.filter(cancellation_invoice=False, order__pharmacy=pharmacy).order_by('-date_time', '-id')

    if 'status' in parameters:
            
        selected_statuses = []

        for item in InvoiceStatus:
            if item[1] in parameters.get('status').split(','):
                selected_statuses.append(item[0])

        invoices = invoices.filter(status__in=selected_statuses)

    if 'search' in parameters:
            
        search_word = parameters.get('search')

        values = search_word.split(' ')
        values = [value for value in values if value]

        q_objects = Q()

        for value in values:

            q_objects &= (
                Q(invoice_number__icontains=value) |
                Q(order__number__icontains=value) |
                Q(order__customer__user__first_name__icontains=value) |
                Q(order__customer__user__last_name__icontains=value) |
                Q(order__customer__user__email__icontains=value)
            )

        invoices = invoices.filter(q_objects).distinct()

    return invoices.distinct()

def create_stock_product_log(stock_product, amount, action, reason, user='System'):
    """ Create stock product log """

    StockProductsLogger.objects.create(
        stock_product = stock_product,
        amount = amount,
        action = action,
        reason = reason,
        user = user,
    )

def create_package_log(package, amount, action, reason, user='System'):
    """ Create package log """

    PackagesLogger.objects.create(
        package = package,
        amount = amount,
        action = action,
        reason = reason,
        user = user,
    )

def get_order_package_data(order_id):
    """ Get packages and products for order to pack """

    response = {}

    try:
        order = Orders.objects.get(id=order_id)
    except:
        response['error'] = 'Order not found'
        return response


    # Get all order products
    order_products = OrderProducts.objects.filter(order=order)

    # Get all avaliable packages with amount > 0
    available_packages = Packages.objects.filter(
        pharmacy=order.pharmacy,
        amount__gt=0,
    ).order_by('-size', 'created_at')

    # Create defaultdict with list als standard value
    aggregated_data = defaultdict(list)

    # Group datas by size of package and append to list
    for package in available_packages:
        aggregated_data[package.size].append({
            'id': package.id,
            'name': package.name,
            'manufacturer': package.manufacturer,
            'batch_number': package.batch_number,
            'amount': package.amount,
        })

    # Get all avaliable packages
    package_list = [{'size': size, 'packages': packages} for size, packages in aggregated_data.items()]

    # Get all avaliable packagesizes
    package_sizes = list(aggregated_data.keys())

    # Create packed products constilations
    packed_products = []

    for order_product in order_products:

        # Get get all stock products for order product
        all_stock_batches = StockProducts.objects.filter(
            product=order_product.product,
            pharmacy=order.pharmacy,
            amount__gte=order_product.amount,
        ).order_by('created_at')

        # List stock batches datas as array
        all_stock_batches_data = list(all_stock_batches.values('id', 'batch_number', 'amount', 'verification_number'))

        # Check if stock batches exist for order_product
        if not all_stock_batches.exists():

            packed_products.append({
                'product': {
                    'productId': order_product.product.id,
                    'orderedProductId': order_product.id,
                    'name': order_product.product.name,
                },
                'packages': [],
                'error': 'Keine ausreichende Menge verfügbar.',
            })

            continue

        # Get latest stock batch
        latest_stock_batch = all_stock_batches.first()
        local_ordered_amount = order_product.amount

        # Dictionary to store the count of each size needed
        needed_package_sizes = {
            size: {'amount': 0, 'fill_amount': []} for size in package_sizes
        }
        needed_package_batches = []

        # Loop through the sizes and use as many as possible
        for i, current_size in enumerate(package_sizes):

            print(local_ordered_amount)
            print(needed_package_batches)
            print(current_size)

            # If the local_ordered_amount is 0, break the loop
            if local_ordered_amount == 0:
                break

            # Get previous and next size available if they exist
            previous_size = package_sizes[i - 1] if i > 0 else None
            next_size = package_sizes[i + 1] if i < len(package_sizes) - 1 else None

            # If the product amount is smaller than the previous (bigger) size
            if previous_size and (local_ordered_amount < int(previous_size) and local_ordered_amount > int(current_size)):
                needed_package_sizes[previous_size]['amount'] += 1
                needed_package_sizes[previous_size]['fill_amount'].append(min(int(previous_size), local_ordered_amount))
                local_ordered_amount -= min(int(previous_size), local_ordered_amount)
            # If the product amount is smaller than the next (smaller) size
            elif next_size and (local_ordered_amount <= int(next_size)):
                continue
            # Standard case local amount fits in current size
            elif local_ordered_amount <= int(current_size):
                needed_package_sizes[current_size]['amount'] += 1
                needed_package_sizes[current_size]['fill_amount'].append(min(int(current_size), local_ordered_amount))
                local_ordered_amount -= min(int(current_size), local_ordered_amount)

            # If local_ordered_amount is bigger than the current size and many packages are needed
            calculated_count = local_ordered_amount // int(current_size)
            needed_package_sizes[current_size]['amount'] += calculated_count
            local_ordered_amount -= calculated_count * int(current_size)

            for _ in range(calculated_count):
                needed_package_sizes[current_size]['fill_amount'].append(int(current_size))

            # Fill the rest with the current size
            if local_ordered_amount > 0 and local_ordered_amount <= int(current_size) and (not next_size or local_ordered_amount > int(next_size)):
                needed_package_sizes[current_size]['amount'] += 1
                needed_package_sizes[current_size]['fill_amount'].append(min(int(current_size), local_ordered_amount))
                local_ordered_amount -= min(int(current_size), local_ordered_amount)

        # Loop through needed_package_sizes and get the packages
        for size, fill_count in needed_package_sizes.items():
            
            # Loop through the amount of packages needed
            for i in range(fill_count['amount']):

                # Loop through the available packages to find the right one
                for package in package_list:

                    # If get size of package is the same as needed size and check if stock_amount is enough
                    # Get first array element to get the latest package
                    if package['size'] == size and package['packages'][0]['amount'] > 0:

                        needed_package_batches.append({
                            'id': package['packages'][0]['id'],
                            'batch_number': package['packages'][0]['batch_number'],
                            'size': size,
                            'fill_amount': fill_count['fill_amount'][i],
                            'unit': 'g' if order_product.product.form == 'flower' else 'Ein.',
                        })
                        # Reduce the amount of the package
                        package['packages'][0]['amount'] -= 1

        # Append the product to the packed_products list
        packed_products.append({
            'product': {
                'productId': order_product.product.id,
                'orderedProductId': order_product.id,
                'name': order_product.product.name,
                'availableBatches': all_stock_batches_data,
                'latestBatch': {
                    'id': latest_stock_batch.id,
                    'batch_number': latest_stock_batch.batch_number,
                    'verificationNumber': latest_stock_batch.verification_number,
                },
            },
            'packages': needed_package_batches,
            'error': 'No packages found' if not needed_package_batches else '',
        })

    return packed_products

######### Send Mails #########
def create_email_log(name, email_type, to_email, from_email, subject, pharmacy=None, message='', sent_success=False):

    email_log = EmailLogger.objects.create(
        name = name,
        email_type = email_type,
        to_email = to_email,
        from_email = from_email,
        pharmacy = pharmacy,
        subject = subject,
        message = message,
        sent_success = sent_success,
    )

    return email_log

def send_register_confirmation(request, user):
    """ Benutzer Registrierungsbestätigung """
    
    main_settings = MainSettings.objects.first()

    if main_settings:

        register_url = get_full_domain(request) + str(reverse('user_activate', kwargs={'uidb64': urlsafe_base64_encode(force_bytes(user.pk)), 'token': account_activation_token.make_token(user)}))

        if main_settings.mail_via_api:
            return True
        else:

            email_settings = EmailSettings.objects.first()

            if email_settings:

                logo_path = get_full_domain(request) + '/static/app/img/mail/logo.png'
                banner_path = get_full_domain(request) + '/static/app/img/mail/titlebanner.png'

                subject = 'Registrierung bei der Dragonweed'

                connection = get_connection(
                    host = email_settings.info_email_host,
                    port = email_settings.info_email_port,
                    username = email_settings.info_email,
                    password = email_settings.info_email_password,
                    use_ssl = True,
                )

                message = render_to_string('mail/registration.html', {
                    'main_settings': main_settings,
                    'user': user,
                    'logo_path': logo_path,
                    'banner_path': banner_path,
                    'register_url': register_url,
                    'domain': get_full_domain(request),
                })

                to_mail_list = [user.email]

                send_mail(subject, 'Registrierung bei der Dragonweed Apotheke', email_settings.info_email, to_mail_list, connection=connection, html_message=message)

def send_order_confirmation(order_id, request):
    """ Send order confirmation to customer and pharmacy """

    order = Orders.objects.get(id=order_id)
    order_products = OrderProducts.objects.filter(order=order)
    recipes = OrderRecipes.objects.filter(order=order)
    main_settings = MainSettings.objects.first()
    pharmacy = order.pharmacy

    if main_settings.mail_via_api:

        brevo_send_order_confirmation(order_id)

    else:

        email_settings = EmailSettings.objects.first()

        if email_settings:

            logo_path = get_full_domain(request) + '/static/app/img/mail/logo.png'
            banner_path = get_full_domain(request) + '/static/app/img/mail/titlebanner.png'


            ##### Bestellbestätigung senden #####
            subject = f'Bestellbestätigung bei { order.pharmacy.name }'

            connection = get_connection(
                host = pharmacy.sending_mail_host,
                port = pharmacy.sending_mail_port,
                username = pharmacy.sending_mail,
                password = pharmacy.sending_mail_password,
                use_ssl = True,
            )

            message = render_to_string('mail/order_confirmation.html', {
                'main_settings': main_settings,
                'user': order.customer.user,
                'logo_path': logo_path,
                'banner_path': banner_path,
                'order': order,
                'pharmacy': order.pharmacy,
                'order_products': order_products,
                'recipe_numbers': ', '.join([item.number for item in recipes]),
                'domain': get_full_domain(request),
            })

            to_mail_list = [order.email_address]

            send_mail(subject, 'Bestellbestätigung', pharmacy.sending_mail, to_mail_list, connection=connection, html_message=message)


            ##### Neue Bestellung an die Apotheke #####
            subject = f'Neue Bestellung eingegangen: Nr. { str(order.number) }'

            connection = get_connection(
                host = pharmacy.sending_mail_host,
                port = pharmacy.sending_mail_port,
                username = pharmacy.sending_mail,
                password = pharmacy.sending_mail_password,
                use_ssl = True,
            )

            message = render_to_string('mail/new_order_to_pharmacy.html', {
                'logo_path': logo_path,
                'banner_path': banner_path,
                'order': order,
                'dashboard_url': get_full_domain(request) + '/dashboard/orders',
            })

            to_mail_list = []

            for recipient in EmailRecipients.objects.filter(category='new_order', pharmacy=order.pharmacy):
                to_mail_list.append(recipient.email)

            send_mail(subject, f'Neue Bestellung eingegangen: Nr. { str(order.number) }', pharmacy.sending_mail, to_mail_list, connection=connection, html_message=message)


            ##### Neue Bestellung an die Apotheke #####
            # if order.recipe_status == 'not_received':

            #     subject = 'Rezept einsenden'

            #     connection = get_connection(
            #         host = email_settings.info_email_host,
            #         port = email_settings.info_email_port,
            #         username = email_settings.info_email,
            #         password = email_settings.info_email_password,
            #         use_ssl = True,
            #     )

            #     message = render_to_string('mail/send_recipe.html', {
            #         'user': order.customer.user,
            #         'logo_path': logo_path,
            #         'banner_path': banner_path,
            #         'order': order,
            #         'recipe_download': get_full_domain(request) + main_settings.prepaid_envelope.url,
            #         'dashboard_url': get_full_domain(request) + '/dashboard/orders',
            #     })

            #     to_mail_list = [order.email_address]

            #     send_mail(subject, 'Rezept einsenden', f'Dragonweed Apotheke <{ email_settings.info_email }>', to_mail_list, connection=connection, html_message=message)

def send_new_order_created(order_id, request):
    """ Send new order to customer """

    main_settings = MainSettings.objects.first()
    order = Orders.objects.get(id=order_id)
    user = order.customer.user
    email_settings = EmailSettings.objects.first()
    pharmacy = order.pharmacy

    url = get_full_domain(request) + str(reverse('confirm_order', kwargs={'order_id': order_id, 'uidb64': urlsafe_base64_encode(force_bytes(user.pk)), 'token': account_activation_token.make_token(user)}))

    if main_settings.mail_via_api:

        brevo_send_new_order_created(order_id, url)

    else:

        if email_settings and order.status in ['started']:

            logo_path = get_full_domain(request) + '/static/app/img/mail/logo.png'
            banner_path = get_full_domain(request) + '/static/app/img/mail/titlebanner.png'

            ##### Neue Bestellung senden #####
            subject = f'Bestellung bei { order.pharmacy.name } bestätigen'

            connection = get_connection(
                host = pharmacy.sending_mail_host,
                port = pharmacy.sending_mail_port,
                username = pharmacy.sending_mail,
                password = pharmacy.sending_mail_password,
                use_ssl = True,
            )

            message = render_to_string('mail/new_order_created.html', {
                'main_settings': MainSettings.objects.first(),
                'user': order.customer.user,
                'logo_path': logo_path,
                'banner_path': banner_path,
                'order': order,
                'pharmacy': order.pharmacy,
                'confirm_order_url': url,
            })

            to_mail_list = [order.email_address]

            send_mail(subject, 'Bestellung bestätigen', order.pharmacy.email, to_mail_list, connection=connection, html_message=message)

def send_invoice_to_customer(invoice_id, request=None):
    """ Send inovice to customer """

    main_settings = MainSettings.objects.first()
    invoice = Invoices.objects.get(id=invoice_id)
    invoice_file = generate_invoice_customer(invoice_id, 'base')
    pharmacy = invoice.order.pharmacy

    if main_settings.mail_via_api:
            
        if invoice.pro_forma_invoice:
            brevo_send_pre_invoice(invoice_id, invoice_file)
        else:
            brevo_send_invoice(invoice_id, invoice_file)

    else:

        email_settings = EmailSettings.objects.first()

        if email_settings:

            logo_path = get_full_domain(request) + '/static/app/img/mail/logo.png'
            banner_path = get_full_domain(request) + '/static/app/img/mail/titlebanner.png'

            ##### Rechnung senden #####
            subject = f'Zahlungsaufforderung und Rechnung - { invoice.invoice_number }' if invoice.pro_forma_invoice else f'Rechnung - { invoice.invoice_number }'

            connection = get_connection(
                host = pharmacy.sending_mail_host,
                port = pharmacy.sending_mail_port,
                username = pharmacy.sending_mail,
                password = pharmacy.sending_mail_password,
                use_ssl = True,
            )

            template = 'mail/pre_invoice_to_customer.html' if invoice.pro_forma_invoice else 'mail/invoice_to_customer.html'

            message = render_to_string(template, {
                'main_settings': main_settings,
                'user': invoice.customer.user,
                'logo_path': logo_path,
                'banner_path': banner_path,
                'invoice': invoice,
                'pharmacy': pharmacy,
                'domain': get_full_domain(request),
            })

            to_mail_list = [invoice.order.email_address]

            email = EmailMessage(
                subject=subject,
                body=message,
                from_email=invoice.order.pharmacy.email,
                to=to_mail_list,
                connection=connection,
            )

            email.content_subtype = 'html'

            # Dekodieren des Base64-Strings in Binärdaten
            pdf_data = base64.b64decode(invoice_file)

            # Anhängen der PDF-Datei
            filename = f'Proformarechnung-{ invoice.invoice_number }.pdf' if invoice.pro_forma_invoice else f'Rechnung-{ invoice.invoice_number }.pdf'
            email.attach(filename, pdf_data, 'application/pdf')

            email.send()

    invoice.send_to_customer = True
    invoice.save()

def send_pre_invoice_to_customer(invoice_id, request):
    """ Send pro forma inovice to customer """

    main_settings = MainSettings.objects.first()

    if main_settings.mail_via_api:
        return True

    else:
        send_invoice_to_customer(invoice_id, request)

def send_receipe_confirmation(order_id):
    """ Send receipe confirmation to customer """

    order = Orders.objects.get(id=order_id)

    return True

def send_reset_password_link(user, request):
    """ User reset password link """
    
    main_settings = MainSettings.objects.all().first()

    if main_settings:

        reset_url = get_full_domain(request) + str(reverse('user_password_reset', kwargs={'uidb64': urlsafe_base64_encode(force_bytes(user.pk)), 'token': account_activation_token.make_token(user)}))

        if main_settings.mail_via_api:
            return True

        else:

            email_settings = EmailSettings.objects.first()

            if email_settings:

                logo_path = get_full_domain(request) + '/static/app/img/mail/logo.png'
                banner_path = get_full_domain(request) + '/static/app/img/mail/titlebanner.png'

                subject = 'Passwort zurücksetzen'

                connection = get_connection(
                    host = email_settings.info_email_host,
                    port = email_settings.info_email_port,
                    username = email_settings.info_email,
                    password = email_settings.info_email_password,
                    use_ssl = True,
                )

                message = render_to_string('mail/reset_password.html', {
                    'main_settings': main_settings,
                    'user': user,
                    'logo_path': logo_path,
                    'banner_path': banner_path,
                    'reset_url': reset_url,
                    'domain': get_full_domain(request),
                })

                to_mail_list = [user.email]

                send_mail(subject, 'Passwort zurücksetzen', email_settings.info_email, to_mail_list, connection=connection, html_message=message)

def send_user_delete_confirmation(user_name, user_email):
    """ Send delete confirmation to user """

    return True

def send_payment_status_received(order, request=None):
    """ Send to customer, that payment is received """

    main_settings = MainSettings.objects.first()
    pharmacy = order.pharmacy

    if main_settings.mail_via_api:
        return True

    else:

        email_settings = EmailSettings.objects.first()

        if email_settings:

            logo_path = get_full_domain(request) + '/static/app/img/mail/logo.png'
            banner_path = get_full_domain(request) + '/static/app/img/mail/titlebanner.png'

            ##### Rechnung senden #####
            subject = f'Zahlung erhalten - Bestell-Nr. { order.number }'

            connection = get_connection(
                host = pharmacy.sending_mail_host,
                port = pharmacy.sending_mail_port,
                username = pharmacy.sending_mail,
                password = pharmacy.sending_mail_password,
                use_ssl = True,
            )

            template = 'mail/payment_received.html'

            message = render_to_string(template, {
                'main_settings': main_settings,
                'user': order.customer.user,
                'logo_path': logo_path,
                'banner_path': banner_path,
                'order': order,
                'pharmacy': pharmacy,
                'domain': get_full_domain(request),
            })

            to_mail_list = [order.email_address]

            email = EmailMessage(
                subject=subject,
                body=message,
                from_email=pharmacy.sending_mail,
                to=to_mail_list,
                connection=connection,
            )

            email.content_subtype = 'html'
            email.send()

    return True

def send_recipe_status_received(order):
    """ Send to customer, that recipe is received """

    return True

def send_recipe_status_incorrect(order):
    """ Send to customer, that recipe is incorrect """

    return True

def send_order_ready_for_pickup(order, request=None):
    """ Send mail to customer, that order is ready for pickup """

    main_settings = MainSettings.objects.first()

    if main_settings.mail_via_api:
        brevo_order_ready_for_pickup(order.id)

def send_order_status_shipped(order):
    """ Send to customer, that order is shipped """

    # brevo_send_order_shipped(order.id)
    main_settings = MainSettings.objects.first()
    pharmacy = order.pharmacy

    if main_settings.mail_via_api:

        brevo_send_order_shipped(order.id)

    else:

        email_log = create_email_log(
            name = f'Bestellung versendet - Bestell-Nr. { order.number }',
            email_type = 'order_shipped',
            to_email = order.email_address,
            from_email = pharmacy.sending_mail,
            pharmacy = pharmacy,
            subject = f'Bestellung versendet - Bestell-Nr. { order.number }',
        )

        logo_path = 'https://pharmaleaf.dragonweed.de/static/app/img/mail/logo.png'
        banner_path = 'https://pharmaleaf.dragonweed.de/static/app/img/mail/titlebanner.png'

        ##### Bestellbestätigung senden #####
        subject = f'Bestellung versendet - Bestell-Nr. { order.number }'

        connection = get_connection(
            host = pharmacy.sending_mail_host,
            port = pharmacy.sending_mail_port,
            username = pharmacy.sending_mail,
            password = pharmacy.sending_mail_password,
            use_ssl = True,
        )

        message = render_to_string('mail/order_shipped.html', {
            'main_settings': main_settings,
            'user': order.customer.user,
            'logo_path': logo_path,
            'banner_path': banner_path,
            'order': order,
            'pharmacy': pharmacy,
            'domain': 'https://dragonweed.de',
        })

        email_log.message = message
        email_log.save()

        to_mail_list = [order.email_address]

        num_send = send_mail(subject, 'Bestellung versendet', pharmacy.sending_mail, to_mail_list, connection=connection, html_message=message)

        print(num_send)

        if num_send > 0:
            email_log.sent_success = True
            email_log.save()

    return True

def send_order_status_cancelled(order):
    """ Send to customer, that order is shipped """

    main_settings = MainSettings.objects.first()

    if main_settings.mail_via_api:
        
        brevo_send_order_cancelled(order.id)

def send_product_request_customer(product_request):
    """ Produktanfrage an Kunden schicken """

    return True

def send_product_request_pharmacy(product_request):
    """ Produktanfrage an Apotheke schicken """

    return True

def send_product_request_approval(request, product_request):
    """ Produktanfrage bestätigt """

    return True

def send_product_request_rejection(product_request):
    """ Produktanfrage abgelehnt """

    return True

def send_payment_reminder(order):
    """ Send payment reminder """

    if order.payment_status in ['pending', 'invoice_reminder'] and \
        order.customer_type == 'self_payer':

        try:
            invoice = Invoices.objects.get(order=order, cancellation_invoice=False, canceled=False)
        except MultipleObjectsReturned:
            invoice = Invoices.objects.filter(order=order, cancellation_invoice=False, canceled=False).last()
        except ObjectDoesNotExist:
            invoice = None

        if invoice:
            
            #  TO-Do: E-Mail Senden

            order.payment_status = 'invoice_reminder'
            order.invoice_reminder_send_on = timezone.now()
            order.save()

            customer = Customers.objects.get(id=order.customer.id)
            customer.blocked = True
            customer.save()

    return True

def send_last_payment_reminder(order):
    """ Send last reminder """

    try:
        invoice = Invoices.objects.get(order=order, cancellation_invoice=False, canceled=False, order_invoice=True)
    except MultipleObjectsReturned:
        invoice = Invoices.objects.filter(order=order, cancellation_invoice=False, canceled=False, order_invoice=True).last()
    except ObjectDoesNotExist:
        invoice = None

        if invoice:

            # TO-Do: E-Mail senden 

            order.payment_status = 'last_reminder'
            order.last_reminder_send_on = timezone.now()
            order.save()

    return True

def send_overdue_mail(order):

    if order.payment_status in ['last_reminder', 'overdue'] and \
        order.customer_type == 'self_payer':

        try:
            invoice = Invoices.objects.get(order=order, cancellation_invoice=False, canceled=False)
        except MultipleObjectsReturned:
            invoice = Invoices.objects.filter(order=order, cancellation_invoice=False, canceled=False).last()
        except ObjectDoesNotExist:
            invoice = None

        if invoice:
            
            # TO-Do: E-Mail senden

            order.payment_status = 'overdue'
            order.save()

    return True

######### Dashboard E-Mails #########
def send_activate_staff_user(user, request):

    reset_url = get_full_domain(request) + str(reverse('dashboard_activate_user', kwargs={'uidb64': urlsafe_base64_encode(force_bytes(user.pk)), 'token': account_activation_token.make_token(user)}))

    brevo_send_activate_staff_user(user, reset_url)

######### System Mails #########
def send_error_mail(log_item):

    main_settings = MainSettings.objects.first()

    subject = f'{ log_item.get_category_display().upper() }'

    connection = get_connection(
                    host = main_settings.error_mail_host,
                    port = main_settings.error_mail_port,
                    username = main_settings.error_mail,
                    password = main_settings.error_mail_password,
                    use_ssl = True,
                )
    
    connection_settings = {
        'host': main_settings.error_mail_host,
        'port': main_settings.error_mail_port,
        'username': main_settings.error_mail,
        'password': main_settings.error_mail_password,
        'use_ssl': True,
    }

    message = f'Message:\n{ log_item.message }\n\nTrace Log:\n{ log_item.stack_trace }'

    to_mail_list = []

    for recipient in EmailRecipients.objects.filter(category='error_message'):
        to_mail_list.append(recipient.email)

    try:
        send_mail(subject, message, main_settings.error_mail, to_mail_list, connection=connection)
    except:

        # Create log entry
        log_entry = {
            'reference': 'Mail Connection error',
            'message': 'E-Mail could not be send',
            'stack_trace': f'Conneciton settings: { connection_settings }',
            'category': 'error'
        }
        create_log(**log_entry)

def get_graph_token():
    tenant_id = '50aaa9fc-d391-47b6-bba5-29f300bb16e5'
    client_id = 'a3991154-0740-47a6-93bb-4edf73e1e2fa'
    client_secret = '-Ac8Q~gEWen3rzUJoNdOdie~OuiOHUb5o13TVat7'
    authority = 'https://login.microsoftonline.com/' + tenant_id
    scope = ['https://graph.microsoft.com/.default']

    client = msal.ConfidentialClientApplication(client_id, authority=authority, client_credential=client_secret)

    token_result = client.acquire_token_silent(scope, account=None)

    if token_result:
        access_token = 'Bearer ' + token_result['access_token']

    if not token_result:
        token_result = client.acquire_token_for_client(scopes=scope)
        access_token = 'Bearer ' + token_result['access_token']

    return(access_token)

def get_products_file():
    token = get_graph_token()

    # TODO: Replace URL
    file_url = 'replaceurl'

    headers = {
        'Authorization': token,
    }

    graph_result = requests.get(url=file_url, headers=headers)

    result = graph_result.json()

    products_file = BytesIO(requests.get(result['@microsoft.graph.downloadUrl']).content)

    return products_file

def shorten_string(input_string, length):
    if len(input_string) > length:
        return input_string[:length] + '...'
    return input_string